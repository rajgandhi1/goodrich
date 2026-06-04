from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from packages.core.document_reader import _csv_to_text, _excel_to_text


def _save_workbook(wb: Workbook) -> bytes:
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_excel_to_text_prefers_enquiry_check_sheet_in_review_workbook():
    wb = Workbook()
    summary = wb.active
    summary.title = "1-Summary Dashboard"
    summary.append(["Metric", "Value"])
    summary.append(["Total Line Items", 2])

    enquiry = wb.create_sheet("2-Enquiry Check")
    enquiry.append([
        "S.No",
        "Customer Enquiry Description",
        "Detected Product Type",
        "GGPL Quote Description",
        "Qty",
        "UoM",
    ])
    enquiry.append([1, "6 INCH CLASS 300 INSULATION KIT", "ISK", "6 x 300# ISK", 4, "KIT"])
    enquiry.append([2, "2 INCH CLASS 150 RTJ GASKET", "RTJ", "R-23 RTJ SS316", 10, "NOS"])

    deviations = wb.create_sheet("3-Deviations Log")
    deviations.append(["S.No", "Issue Type", "Customer Said", "Action Required", "Severity"])
    deviations.append([1, "Missing class", "size only", "Confirm rating", "RED"])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 2
    assert "=== Sheet: 2-Enquiry Check ===" in text
    assert "=== Sheet: 1-Summary Dashboard ===" not in text
    assert "=== Sheet: 3-Deviations Log ===" not in text
    assert "6 INCH CLASS 300 INSULATION KIT" in text
    assert "Missing class" not in text


def test_excel_to_text_keeps_plain_customer_sheet_without_review_tabs():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "RFQ"
    sheet.append(["Description", "Qty", "UoM"])
    sheet.append(["SPIRAL WOUND GASKET 4 INCH 150#", 8, "NOS"])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 1
    assert "=== Sheet: RFQ ===" in text
    assert "SPIRAL WOUND GASKET 4 INCH 150#" in text


def test_excel_to_text_merges_stacked_customer_headers():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(["RFQ_GASKETS"])
    sheet.append(["SL NO", "ITEM DESCRIPTION", "PIPE CLASS", "DESIGN DATA", "", "", "STANDARDS", "QTY WITH MARGIN", "UOM"])
    sheet.append(["", "", "", "SIZE1 (inch)", "CLASS", "FACING", "MATERIAL AND DIMENSIONAL STANDARD", "", ""])
    sheet.append([221, "Gasket, Flat", "B", "2", "150", "RF", "GASKET, 1.5MM, CNAF, ASME B16.21", 73, "Nos."])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 1
    assert "SIZE1 (inch)" in text
    assert "CLASS" in text
    assert "MATERIAL AND DIMENSIONAL STANDARD" in text
    assert "| Sheet1 | 4 | 1 | 221 | Gasket, Flat | B | 2 | 150 | RF |" in text


def test_excel_to_text_drops_footer_and_total_rows():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "RFQ"
    sheet.append(["SR. NO.", "MOC", "TYPE", "DIMESION/RING SIZE", "QTY"])
    sheet.append([1, "Modified PTFE", "Flat Ring", '2", CL 150, 1.5MM THK, B16.21', 4])
    sheet.append(["Additional Requirements-", "", "", "", ""])
    sheet.append(["1)", "All material to be supplied with mill certificates", "", "", ""])
    sheet.append(["", "", "Total", "", 4])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 1
    assert "Modified PTFE" in text
    assert "Additional Requirements" not in text
    assert "mill certificates" not in text
    assert "Total" not in text


def test_excel_to_text_marks_bare_numeric_size_columns_as_inches():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "BOM"
    sheet.append(["Size", "Description", "Qty to be purchased", "UOM"])
    sheet.append([3, "PTFE FLAT RING GASKET,1.6MM THK B16.21/16.5", 120, "PC"])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 1
    assert '| BOM | 2 | 1 | 3" | PTFE FLAT RING' in text


def test_csv_to_text_repairs_reference_csv_with_description_in_header():
    source = (
        '"GASKET,FLANGE NONSPIRAL;FULL FACE;20"" PIPE;150LB;TEFLON;QA/QC CERT REQ A,C,D",'
        'Customer,Reference,S No,Discription,GGPL DESCRIPTPION,Deviations,TYPE OF PRODUCT,NON STANDARD\n'
        ',L&T,RFQ-1,2,"NPS 2, Gasket Spiral wound, SS316 with flexible graphite filler",'
        '"SIZE : 2"" X 150# X 4.5MM THK ,SS316 SPIRAL WOUND GASKET",,Spiral Wound,\n'
    )

    text, truncated, row_count = _csv_to_text(source)

    assert truncated is False
    assert row_count == 2
    assert "=== Sheet: CSV ===" in text
    assert "Customer Enquiry Description" in text
    assert "| CSV | 1 | 1 |" in text
    assert "| CSV | 2 | 2 |" in text
    assert 'GASKET,FLANGE NONSPIRAL' in text
    assert "NPS 2, Gasket Spiral wound" in text
    assert "GGPL Quote Description" in text


def test_csv_to_text_handles_plain_header_csv():
    source = (
        "Customer Enquiry Description,Qty,UoM\n"
        "\"RTJ Ring number - 23 , Moc :- Inconel 825 ,Hardness required - 160 HBN\",4,NOS\n"
    )

    text, truncated, row_count = _csv_to_text(source.encode("utf-8"))

    assert truncated is False
    assert row_count == 1
    assert "RTJ Ring number - 23" in text
    assert "| CSV | 2 | 1 |" in text
