from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pdfplumber

from core.quote_exporter import build_quotation_excel
from core.quote_pdf import build_quotation_pdf

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.modules.pop("app", None)

from app.services.export_parity import compare_excel, compare_pdf_text, extract_pdf_text
from app.services.export_service import _logo_path, build_pdf, build_xlsx


def _sample_items() -> list[dict]:
    return [
        {
            "line_no": 1,
            "customer_sl_no": "C-10",
            "customer_item_code": "ITEM-777",
            "quantity": 2,
            "uom": "NOS",
            "raw_description": '4" 150# CNAF RF gasket 3mm ASME B16.21',
            "ggpl_description": 'GASKET, SIZE : 4", RATING : 150#, TYPE : SOFT CUT, MOC : CNAF, THK : 3 MM',
            "gasket_type": "SOFT_CUT",
            "moc": "CNAF",
            "rating": "150#",
            "status": "ready",
        },
        {
            "line_no": 2,
            "customer_sl_no": "C-11",
            "customer_item_code": "ITEM-888",
            "quantity": 1,
            "uom": "NOS",
            "raw_description": '8" 300# RTJ R45 SS316',
            "ggpl_description": "REGRET - CANNOT PRODUCE",
            "gasket_type": "RTJ",
            "moc": "SS316",
            "rating": "300#",
            "status": "regret",
        },
    ]


def _sample_quote_data(currency: str = "INR") -> dict:
    return {
        "quote_no": f"PARITY-{currency}",
        "quote_date": "16 May 2026",
        "rev_no": "0",
        "rev_date": "",
        "include_customer_sl_no": True,
        "include_customer_item_code": True,
        "buyer_name_address": "ACME Industries\nMumbai",
        "buyer_name": "ACME Industries",
        "buyer_address_line1": "Plot 10, Industrial Area",
        "buyer_address_line2": "Andheri East",
        "buyer_city": "Mumbai",
        "buyer_state": "Maharashtra",
        "buyer_pin_code": "400093",
        "buyer_country": "India",
        "customer_enq_no": "ENQ-123",
        "attention": "Procurement",
        "designation": "Manager",
        "contact_no": "9999999999",
        "telephone_no": "022-40000000",
        "email": "buyer@example.com",
        "rep_name": "GGPL Sales",
        "rep_designation": "Sales",
        "rep_contact": "8888888888",
        "rep_email": "sales@goodrichgasket.com",
        "currency": currency,
        "fx_rate": 83 if currency == "USD" else 1,
        "unit_prices": [1000, 0],
        "discount_pct": 5,
        "gst_type": "IGST",
        "gst_pct": 18,
        "price_basis": "FOR BASIS",
        "validity_days": "7",
        "packing": "INCLUSIVE",
        "freight": "INCLUSIVE",
        "payment_terms": "30% ADVANCE & 70% BALANCE BEFORE DISPATCH OF MATERIAL",
        "bank_charges": "TO YOUR ACCOUNT",
        "delivery": "2 weeks",
        "inspection": "Not Applicable",
        "insurance": "TO YOUR ACCOUNT",
        "hsn_code": "84841010",
        "ld_clause": "Not Applicable",
        "cancellation": "Products are manufactured on order.",
        "min_order_value": "Minimum Order Value is INR 10,000.",
        "technical_deviation_remarks": "Technical deviation accepted as per customer drawing.",
        "commercial_tnc": "Commercial terms remain subject to final PO acceptance.",
        "technical_notes": "Technical note line one.\nTechnical note line two.",
    }


def test_excel_export_service_matches_current_python_exporter():
    items = _sample_items()
    quote_data = _sample_quote_data()
    direct = build_quotation_excel(items, quote_data, logo_path=_logo_path())
    wrapped = build_xlsx(items, quote_data)
    diffs = compare_excel(direct, wrapped)
    assert diffs == []
    workbook = openpyxl.load_workbook(io.BytesIO(direct), data_only=True)
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
    assert "C-10" in values
    assert "ITEM-777" in values
    assert "Technical deviation accepted as per customer drawing." in values
    assert not any(cell == "Commercial terms remain subject to final PO acceptance." for cell in values)


def test_blank_technical_notes_use_default_export_text():
    items = _sample_items()
    quote_data = {**_sample_quote_data(), "technical_notes": ""}
    expected_certification = "1. Certifications: MTC to EN10204-3.1 for metallic parts and EN10204-2.1 for non-metallic."
    expected_testing = "2. Testing Charges for gasket will be extra at actuals for tests other than compression & sealability test and chemical analysis."

    workbook = openpyxl.load_workbook(io.BytesIO(build_quotation_excel(items, quote_data, logo_path=_logo_path())), data_only=True)
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
    assert f"{expected_certification}\n{expected_testing}" in values

    text = extract_pdf_text(build_quotation_pdf(items, quote_data, logo_path=_logo_path()))
    assert expected_certification in text
    assert expected_testing in text


def test_pdf_continuation_pages_skip_buyer_block_and_final_items_show_totals():
    items = []
    for index in range(35):
        items.append(
            {
                **_sample_items()[0],
                "line_no": index + 1,
                "customer_sl_no": f"C-{index + 1}",
                "customer_item_code": f"ITEM-{index + 1}",
                "quantity": 2,
                "raw_description": f'4" 150# CNAF RF gasket line {index + 1} with customer reference',
            }
        )
    quote_data = {**_sample_quote_data(), "unit_prices": [100] * len(items)}

    direct = build_quotation_pdf(items, quote_data, logo_path=_logo_path())
    text = extract_pdf_text(direct)
    assert "Total Quantity" in text
    assert "Total Price (INR)" in text
    assert "7000.00" in text
    assert "Tax Breakup" in text
    assert "IGST @ 18%" in text
    assert "Total Combined Price (INR)" in text
    assert "7847.00" in text
    with pdfplumber.open(io.BytesIO(direct)) as pdf:
        assert len(pdf.pages) > 3
        second_page_text = pdf.pages[1].extract_text() or ""
        assert "SALES QUOTATION" in second_page_text
        assert "Name & Address of the Buyer" not in second_page_text
        assert "Customer Enq No" not in second_page_text
        totals_page = next(page for page in pdf.pages if "Total Quantity" in (page.extract_text() or ""))
        totals_words = totals_page.extract_words(x_tolerance=1, y_tolerance=3)
        total_quantity_word = next(
            word
            for word in totals_words
            if word["text"] == "Quantity" and word["x0"] > 345
        )
        assert total_quantity_word["x0"] >= 345


def test_pdf_export_service_text_matches_current_python_exporter():
    items = _sample_items()
    quote_data = _sample_quote_data("USD")
    direct = build_quotation_pdf(items, quote_data, logo_path=_logo_path())
    wrapped = build_pdf(items, quote_data)
    matched, expected, actual = compare_pdf_text(direct, wrapped)
    assert matched, f"Expected PDF text:\n{expected}\n\nActual PDF text:\n{actual}"
    text = extract_pdf_text(direct)
    assert "SALES QUOTATION" in text
    assert "QUOTATION NO.:" in text
    assert "*Please refer to the email" not in text
    assert "Name : ACME Industries" in text
    assert "City/State/PIN : Mumbai, Maharashtra, 400093" in text
    assert "C-10" in text
    assert "ITEM-777" in text
    assert "Technical Deviation / Remarks:" in text
    assert "Technical deviation accepted as per customer drawing." in text
    assert "Commercial T&C:" not in text
    assert "Commercial terms remain subject to final PO acceptance." not in text
    assert text.index("Technical Deviation / Remarks:") < text.index("Other Terms & Conditions")
    assert text.index("Other Terms & Conditions") < text.index("Technical Notes")
    assert text.index("Other Terms & Conditions") < text.index("GENERAL TERMS OF QUOTATION:")
    assert "For Goodrich Gaskets" in text
    assert "Authorized Signatory and Company Seal" in text
    assert "Accepted By -" in text
    assert "Client Name :" in text
    assert "This is a Computer Generated Document" not in text
    with pdfplumber.open(io.BytesIO(direct)) as pdf:
        first_page = pdf.pages[0]
        words = first_page.extract_words(x_tolerance=1, y_tolerance=3)
        buyer_email = next(word for word in words if word["text"] == "buyer@example.com")
        buyer_bottom_lines = [
            line["top"]
            for line in first_page.lines
            if 357.8 <= line.get("top", 0) <= 358.5
        ]
        assert buyer_bottom_lines
        assert buyer_email["bottom"] < min(buyer_bottom_lines)
        telephone_label = next(
            word
            for word in words
            if word["text"] == "Telephone" and word["x0"] < 30 and 320 <= word["top"] <= 340
        )
        customer_email_label = next(
            word
            for word in words
            if word["text"] == "Email" and word["x0"] < 30 and 335 <= word["top"] <= 355
        )
        assert customer_email_label["top"] - telephone_label["top"] >= 14
        other_page = next(page for page in pdf.pages if "Other Terms & Conditions" in (page.extract_text() or ""))
        other_words = other_page.extract_words(x_tolerance=1, y_tolerance=3)
        price_word = next(word for word in other_words if word["text"] == "Price")
        price_colon = next(
            word
            for word in other_words
            if word["text"] == ":" and abs(word["top"] - price_word["top"]) < 1
        )
        price_value = next(
            word
            for word in other_words
            if word["text"] == "FOR" and abs(word["top"] - price_word["top"]) < 1
        )
        assert 46 <= price_word["x0"] <= 49
        assert 176 <= price_colon["x0"] <= 179
        assert 181 <= price_value["x0"] <= 184
