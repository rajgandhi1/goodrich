from __future__ import annotations
"""
Loads and exposes reference data from Excel files at module level.
All lookups are keyed on normalized (size, rating) tuples.
"""
import os
import openpyxl

_REF_DIR = os.path.join(os.path.dirname(__file__), '..', 'reference')

# NB (mm) → NPS (inch string)
NB_TO_NPS = {
    15: '0.50"', 20: '0.75"', 25: '1"', 32: '1.25"', 40: '1.5"',
    50: '2"', 65: '2.5"', 80: '3"', 90: '3.50"', 100: '4"',
    125: '5"', 150: '6"', 200: '8"', 250: '10"', 300: '12"',
    350: '14"', 400: '16"', 450: '18"', 500: '20"', 550: '22"',
    600: '24"', 650: '26"', 700: '28"', 750: '30"', 800: '32"',
    850: '34"', 900: '36"', 950: '38"', 1000: '40"',
}

# Normalize size string to match reference table keys (e.g. '1"', '0.50"', '1.25"')
def normalize_size(raw: str) -> str | None:
    if not raw:
        return None
    import re as _re
    s = str(raw).strip().upper()
    # Mixed fractions like "1 1/2", "1-1/2", "1 1/2\"" — check BEFORE removing spaces
    # Strip trailing inch/quote chars first so the fraction regex still matches
    s_noquote = _re.sub(r'["\'″]+$', '', s).strip()
    mf = _re.match(r'^(\d+)\s*[-\s]\s*(\d+)/(\d+)\s*$', s_noquote)
    if mf:
        whole, num, den = int(mf.group(1)), int(mf.group(2)), int(mf.group(3))
        s = str(whole + num / den)
    else:
        s = s.replace(' ', '')
        # Remove units
        s = s.replace('NPS', '').replace('DN', '').replace('"', '').replace("'", '').replace('INCH', '').strip()

        # Explicit mm suffix (e.g. "150MM", "125MM") — treat as nominal bore with nearest-NB lookup
        mm_match = _re.match(r'^(\d+(?:\.\d+)?)MM$', s)
        if mm_match:
            return _nb_mm_to_nps(float(mm_match.group(1)))

        # NB lookup: strip NB token, then try exact NB_TO_NPS match.
        # Also handles "DN 150" (DN stripped above → "150") and bare numbers after stripping.
        # Explicit NB label → also try nearest-neighbour fallback.
        has_nb_label = 'NB' in s
        nb_str = s.replace('NB', '').strip()
        try:
            nb = int(float(nb_str))
            if nb in NB_TO_NPS:
                return NB_TO_NPS[nb]
            if has_nb_label:
                # Non-standard NB value — round to nearest
                return _nb_mm_to_nps(float(nb_str))
        except ValueError:
            pass

        # Simple fraction handling
        fractions = {'1/2': '0.50', '3/4': '0.75', '1/4': '0.25', '3/8': '0.375'}
        for frac, dec in fractions.items():
            s = s.replace(frac, dec)
    try:
        val = float(s)
    except ValueError:
        return None
    # Map to standard key format used in reference sheet
    size_map = {
        0.5: '0.50"', 0.75: '0.75"', 1.0: '1"', 1.25: '1.25"', 1.5: '1.5"',
        2.0: '2"', 2.5: '2.5"', 3.0: '3"', 3.5: '3.50"', 4.0: '4"',
        5.0: '5"', 6.0: '6"', 8.0: '8"', 10.0: '10"', 12.0: '12"',
        14.0: '14"', 16.0: '16"', 18.0: '18"', 20.0: '20"', 22.0: '22"',
        24.0: '24"', 26.0: '26"', 28.0: '28"', 30.0: '30"', 32.0: '32"',
        34.0: '34"', 36.0: '36"', 38.0: '38"', 40.0: '40"', 42.0: '42"',
        44.0: '44"', 46.0: '46"', 48.0: '48"', 50.0: '50"', 52.0: '52"',
        54.0: '54"',
    }
    return size_map.get(round(val, 2))


def _nb_mm_to_nps(mm_val: float) -> str | None:
    """Convert a nominal bore mm value to the same-or-next-lower NPS inch string.

    Returns the NB_TO_NPS entry for an exact match, or the next lower standard
    NB value when the input doesn't match exactly (e.g. 127mm → 125 NB → 5").
    Returns None if mm_val is outside a sensible NB range (< 6 or > 1200).
    """
    if mm_val < 6 or mm_val > 1200:
        return None
    nb = int(round(mm_val))
    if nb in NB_TO_NPS:
        return NB_TO_NPS[nb]
    lower_sizes = [k for k in NB_TO_NPS if k <= mm_val]
    if not lower_sizes:
        return None
    return NB_TO_NPS[max(lower_sizes)]


def normalize_rating(raw: str) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().upper().replace(' ', '').replace('#', '').replace('LB', '').replace('CLASS', '').replace('CL', '')
    pn_map = {'PN10': 'PN 10', 'PN16': 'PN 16', 'PN25': 'PN 25', 'PN40': 'PN 40', 'PN6': 'PN 6'}
    if s in pn_map:
        return pn_map[s]
    if s.startswith('PN'):
        return 'PN ' + s[2:]
    asme_map = {'150': '150 #', '300': '300 #', '600': '600 #', '900': '900 #', '1500': '1500 #', '2500': '2500 #'}
    return asme_map.get(s)


def _norm_key(size) -> str:
    """Normalize size string from Excel (may have Unicode curly quotes) to ASCII."""
    return str(size).strip().replace('\u201c', '"').replace('\u201d', '"').replace('\u2019', '"').replace("''", '"')


def _load_dimensions():
    path = os.path.join(_REF_DIR, 'SOFT CUT ASME STANDARD DIMENSION & SOFT CUT MOC.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    rf, ff, moc_list = {}, {}, []
    # Raised face sheet
    ws = wb['SC-RAISED FACE']
    for row in ws.iter_rows(min_row=2, values_only=True):
        size, rating, od, id_ = row[1], row[2], row[3], row[4]
        if size and rating and isinstance(od, (int, float)) and isinstance(id_, (int, float)):
            rf[(_norm_key(size), str(rating).strip())] = {'od': od, 'id': id_}
    # Full face sheet
    ws = wb['SC-FULL FACE']
    for row in ws.iter_rows(min_row=2, values_only=True):
        size, rating, od, id_, pcd, hole_d, holes = row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        if size and rating and isinstance(od, (int, float)) and isinstance(id_, (int, float)):
            ff[(_norm_key(size), str(rating).strip())] = {
                'od': od, 'id': id_, 'pcd': pcd, 'hole_dia': hole_d, 'holes': holes
            }
    # MOC sheet
    ws = wb['SOFT CUT MOC']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            moc_list.append(str(row[1]).strip().upper())
    return rf, ff, moc_list


def _load_examples():
    path = os.path.join(_REF_DIR, 'SOFTCUT DESCRIPTION.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    examples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:
            examples.append({'input': str(row[1]).strip(), 'output': str(row[2]).strip()})
    return examples


RAISED_FACE_DIMS, FULL_FACE_DIMS, ACCEPTED_MOC = _load_dimensions()
FEW_SHOT_EXAMPLES = _load_examples()


def lookup_dimensions(size_norm: str, rating_norm: str, face_type: str) -> "dict | None":
    """Look up OD/ID from reference tables using normalized size and rating keys."""
    if not size_norm or not rating_norm:
        return None
    # Try to find matching key — reference sheet uses formats like '150 #', '300 #'
    table = FULL_FACE_DIMS if face_type == 'FF' else RAISED_FACE_DIMS
    key = (size_norm, rating_norm)
    return table.get(key)


# ---------------------------------------------------------------------------
# RTJ ring number lookup (ASME B16.20 / B16.5 flanges)
# Keys: (normalized_size, class_str) → ring designation string
# Values verified against real GGPL quote data (CSV export).
# Entries without data evidence are omitted — flag as MISSING for manual entry.
# ---------------------------------------------------------------------------
RTJ_RING_TABLE: dict[tuple[str, str], str] = {
    # Source: ASME B16.20 via ASME B16.21 reference table (Customer Enq - Quote Data - ASME.csv)
    # Columns: (NPS size, pressure class) → ring designation
    # "300#–600#" in standard means 300 and 600 share the same ring number.

    # NPS 1/2"
    ('0.50"', '300'): 'R-11', ('0.50"', '600'): 'R-11',
    ('0.50"', '900'): 'R-12', ('0.50"', '1500'): 'R-12', ('0.50"', '2500'): 'R-13',
    # NPS 3/4"
    ('0.75"', '300'): 'R-13', ('0.75"', '600'): 'R-13',
    ('0.75"', '900'): 'R-14', ('0.75"', '1500'): 'R-14', ('0.75"', '2500'): 'R-16',
    # NPS 1"
    ('1"', '150'): 'R-15',
    ('1"', '300'): 'R-16', ('1"', '600'): 'R-16', ('1"', '900'): 'R-16', ('1"', '1500'): 'R-16',
    ('1"', '2500'): 'R-18',
    # NPS 1-1/4"
    ('1.25"', '150'): 'R-17',
    ('1.25"', '300'): 'R-18', ('1.25"', '600'): 'R-18', ('1.25"', '900'): 'R-18', ('1.25"', '1500'): 'R-18',
    ('1.25"', '2500'): 'R-21',
    # NPS 1-1/2"
    ('1.5"', '150'): 'R-19',
    ('1.5"', '300'): 'R-20', ('1.5"', '600'): 'R-20', ('1.5"', '900'): 'R-20', ('1.5"', '1500'): 'R-20',
    ('1.5"', '2500'): 'R-23',
    # NPS 2"
    ('2"', '150'): 'R-22',
    ('2"', '300'): 'R-23', ('2"', '600'): 'R-23',
    ('2"', '900'): 'R-24', ('2"', '1500'): 'R-24',
    ('2"', '2500'): 'R-26',
    # NPS 2-1/2"
    ('2.5"', '150'): 'R-25',
    ('2.5"', '300'): 'R-26', ('2.5"', '600'): 'R-26',
    ('2.5"', '900'): 'R-27', ('2.5"', '1500'): 'R-27',
    ('2.5"', '2500'): 'R-28',
    # NPS 3"
    ('3"', '150'): 'R-29',
    ('3"', '300'): 'R-31', ('3"', '600'): 'R-31', ('3"', '900'): 'R-31',
    ('3"', '1500'): 'R-35', ('3"', '2500'): 'R-32',
    # NPS 3-1/2"
    ('3.50"', '150'): 'R-33',
    ('3.50"', '300'): 'R-34', ('3.50"', '600'): 'R-34', ('3.50"', '900'): 'R-34',
    # NPS 4"
    ('4"', '150'): 'R-36',
    ('4"', '300'): 'R-37', ('4"', '600'): 'R-37', ('4"', '900'): 'R-37',
    ('4"', '1500'): 'R-39', ('4"', '2500'): 'R-38',
    # NPS 5"
    ('5"', '150'): 'R-40',
    ('5"', '300'): 'R-41', ('5"', '600'): 'R-41', ('5"', '900'): 'R-41',
    ('5"', '1500'): 'R-44', ('5"', '2500'): 'R-42',
    # NPS 6"
    ('6"', '150'): 'R-43',
    ('6"', '300'): 'R-45', ('6"', '600'): 'R-45', ('6"', '900'): 'R-45',
    ('6"', '1500'): 'R-46', ('6"', '2500'): 'R-47',
    # NPS 8"
    ('8"', '150'): 'R-48',
    ('8"', '300'): 'R-49', ('8"', '600'): 'R-49', ('8"', '900'): 'R-49',
    ('8"', '1500'): 'R-50', ('8"', '2500'): 'R-51',
    # NPS 10"
    ('10"', '150'): 'R-52',
    ('10"', '300'): 'R-53', ('10"', '600'): 'R-53', ('10"', '900'): 'R-53',
    ('10"', '1500'): 'R-54', ('10"', '2500'): 'R-55',
    # NPS 12"
    ('12"', '150'): 'R-56',
    ('12"', '300'): 'R-57', ('12"', '600'): 'R-57', ('12"', '900'): 'R-57',
    ('12"', '1500'): 'R-58', ('12"', '2500'): 'R-60',
    # NPS 14"
    ('14"', '150'): 'R-59',
    ('14"', '300'): 'R-61', ('14"', '600'): 'R-61',
    ('14"', '900'): 'R-62', ('14"', '1500'): 'R-63',
    # NPS 16"
    ('16"', '150'): 'R-64',
    ('16"', '300'): 'R-65', ('16"', '600'): 'R-65',
    ('16"', '900'): 'R-66', ('16"', '1500'): 'R-67',
    # NPS 18"
    ('18"', '150'): 'R-68',
    ('18"', '300'): 'R-69', ('18"', '600'): 'R-69',
    ('18"', '900'): 'R-70', ('18"', '1500'): 'R-71',
    # NPS 20"
    ('20"', '150'): 'R-72',
    ('20"', '300'): 'R-73', ('20"', '600'): 'R-73',
    ('20"', '900'): 'R-74', ('20"', '1500'): 'R-75',
    # NPS 22" — B16.47 Series A only (no B16.5 ring defined)
    ('22"', '150'): 'R-80',
    ('22"', '300'): 'R-81', ('22"', '600'): 'R-81',
    # NPS 24"
    ('24"', '150'): 'R-76',
    ('24"', '300'): 'R-77', ('24"', '600'): 'R-77',
    ('24"', '900'): 'R-78', ('24"', '1500'): 'R-79',

    # ASME B16.47 Series A (NPS 26" and above)
    # NOTE: for 16" 900# B16.47A ring is R-68 (vs R-66 for B16.5); distinguish by standard if needed.
    ('26"', '300'): 'R-93', ('26"', '600'): 'R-93', ('26"', '900'): 'R-100',
    ('28"', '300'): 'R-94', ('28"', '600'): 'R-94', ('28"', '900'): 'R-101',
    ('30"', '300'): 'R-95', ('30"', '600'): 'R-95', ('30"', '900'): 'R-102',
    ('32"', '300'): 'R-96', ('32"', '600'): 'R-96', ('32"', '900'): 'R-103',
    ('34"', '300'): 'R-97', ('34"', '600'): 'R-97', ('34"', '900'): 'R-104',
    ('36"', '300'): 'R-98', ('36"', '600'): 'R-98', ('36"', '900'): 'R-105',
}


def lookup_rtj_ring(size_norm: str, rating_raw: str) -> str | None:
    """Return ASME B16.20 ring designation for given NPS size and class, or None if not in table."""
    if not size_norm or not rating_raw:
        return None
    import re as _re
    m = _re.search(r'(\d+)', str(rating_raw).replace('#', '').strip())
    cls = m.group(1) if m else None
    if not cls:
        return None
    return RTJ_RING_TABLE.get((size_norm, cls))


def select_few_shot_examples(description: str, n: int = 8) -> list[dict]:
    """Select the most relevant few-shot examples using keyword overlap."""
    desc_words = set(description.upper().split())
    scored = []
    for ex in FEW_SHOT_EXAMPLES:
        ex_words = set(ex['input'].upper().split())
        score = len(desc_words & ex_words)
        scored.append((score, ex))
    scored.sort(key=lambda x: x[0], reverse=True)
    # Always include at least 2 diverse low-overlap examples
    top = [ex for _, ex in scored[:n]]
    return top
