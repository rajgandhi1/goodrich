from openpyxl import load_workbook

paths = [
    r"C:\Users\Raj Gandhi\Downloads\10011364-RFQ for Gasket.xlsx",
    r"C:\Users\Raj Gandhi\Downloads\SO 71409 BOM Gaskets R1.xlsx",
    r"C:\Users\Raj Gandhi\Downloads\Gasket -BOQ (1).xlsx",
    r"C:\Users\Raj Gandhi\Downloads\EXPORT PRICING - AXIOM RFQ-KA-25125.xlsx",
    r"C:\Users\Raj Gandhi\Downloads\GGPL - SPW (KA-25125) Customer spec.xlsx",
    r"C:\Users\Raj Gandhi\Downloads\Gasket -BOQ.xlsx",
]

for path in paths:
    print(f"\n### {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    for ws in wb.worksheets[:4]:
        print(f"SHEET {ws.title!r}: {ws.max_row} rows x {ws.max_column} cols")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True):
            print(" ", row)
