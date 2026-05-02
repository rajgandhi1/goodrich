from __future__ import annotations
"""
Parses customer enquiry inputs (email text or Excel) into a uniform list of raw items:
[{'line_no': int, 'description': str, 'quantity': float, 'uom': str}]
"""
import re
import io
import json
import logging
import openpyxl

logger = logging.getLogger(__name__)


def worksheet_rows_with_merged_values(ws, max_row: int | None = None) -> list[tuple]:
    """Return worksheet rows with merged cells expanded to every covered row.

    openpyxl exposes only the top-left value of a merged range. Enquiry sheets
    often merge MOC/spec cells vertically, so reading values directly makes the
    following line items look incomplete.
    """
    row_limit = min(max_row or ws.max_row, ws.max_row)
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, row_limit + 1)
    ]

    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            continue
        for r in range(merged_range.min_row, min(merged_range.max_row, row_limit) + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                rows[r - 1][c - 1] = value

    return [tuple(row) for row in rows]


def parse_email_text(text: str) -> list[dict]:
    """Extract line items from pasted email body text."""
    items = []
    lines = _merge_continuation_lines([l.strip() for l in text.splitlines() if l.strip()])
    for line in lines:
        item = _parse_line(line)
        if item:
            items.append(item)
    # Number items if line_no missing
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
    return items


def _merge_continuation_lines(lines: list[str]) -> list[str]:
    """Merge lines that are mid-sentence continuations of the previous line.

    Handles cases where a long gasket description wraps across two lines,
    e.g. Excel cell content copied as text where a description ends with
    'OUTER RING' and the next line starts with 'MATERIAL: Carbon Steel...'.
    """
    merged = []
    for line in lines:
        if not merged:
            merged.append(line)
            continue
        prev = merged[-1]
        # A line is a continuation if:
        # 1. It does NOT start with a digit (new serial number), AND
        # 2. The previous line ended with a comma, OR ended with RING/MATERIAL/OUTER/INNER
        #    (cut mid-phrase), OR the current line starts with a field-label ("WORD:") pattern
        starts_with_number = bool(re.match(r'^\d', line))
        # Lines starting with a gasket size-prefix keyword are always new items
        is_new_item_prefix = bool(re.match(r'^(?:NPS|NB|DN|SIZE)\s*:', line, re.IGNORECASE))
        prev_ends_mid = (
            prev.endswith(',')
            or re.search(r'\b(?:OUTER\s*RING|INNER\s*RING|OUTER|RING|MATERIAL)\s*$', prev, re.IGNORECASE)
        )
        # Previous line ends with a standards-body acronym — next line is its number (e.g. "ANSI\n B16.47")
        prev_ends_std_prefix = bool(
            re.search(r'\b(?:ANSI|ASME|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\s*$', prev, re.IGNORECASE)
        )
        curr_is_field_continuation = bool(re.match(r'^[A-Z][A-Z\s]+:\s*\S', line))
        # Current line is a standard reference line (e.g. "ASME B16.20: Metallic Gaskets for Pipe Flanges")
        # that belongs to the preceding description, not a new line item
        curr_is_standard_ref = bool(re.match(
            r'^(?:ASME|ANSI|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\b', line, re.IGNORECASE
        ))
        if (not starts_with_number and not is_new_item_prefix
                and (prev_ends_mid or prev_ends_std_prefix or curr_is_field_continuation or curr_is_standard_ref)):
            merged[-1] = prev + ' ' + line
        else:
            merged.append(line)
    return merged


def _parse_line(line: str) -> "dict | None":
    """Try to extract (description, qty, uom) from a single text line."""
    # Skip header-like lines
    lower = line.lower()
    if any(kw in lower for kw in ['sl.no', 'sl no', 'line no', 'release no', 'notes', 'inv uom',
                                   'description', 'subject', 'dear', 'regards', 'kindly',
                                   'total', 'terms', 'price', 'revision', 'date']):
        return None

    # Pattern: optional number, description text, number, unit
    # e.g. "1  Gasket - Rubber - 6'' PN10  27  m"
    pattern = re.compile(
        r'^(\d+)?\s*'                            # optional sl.no
        r'(?:\d+\s+\d+\s+)?'                     # optional line/release no columns
        r'(.+?)\s+'                              # description (greedy)
        r'(\d+(?:\.\d+)?)\s*'                    # quantity
        r'(nos?|m|mtr|meters?|pcs?|sets?|kgs?|units?)\s*$',  # unit
        re.IGNORECASE
    )
    m = pattern.match(line)
    if m:
        sl_no = int(m.group(1)) if m.group(1) else None
        desc = m.group(2).strip()
        qty = float(m.group(3))
        uom = _normalize_uom(m.group(4))
        if _looks_like_gasket(desc) and qty > 0:
            return {'line_no': sl_no, 'description': desc, 'quantity': qty, 'uom': uom}

    # Simpler fallback: tab/multiple-space separated columns
    parts = re.split(r'\t|  +', line)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 3:
        # Try last two parts as qty + uom, rest as description
        try:
            qty = float(parts[-2])
            uom = _normalize_uom(parts[-1])
            desc = ' '.join(parts[:-2]).strip()
            desc = re.sub(r'^\d+\s+', '', desc).strip()
            if _looks_like_gasket(desc) and qty > 0:
                return {'line_no': None, 'description': desc, 'quantity': qty, 'uom': uom}
        except (ValueError, IndexError):
            pass

    # Last resort: entire line is a description with no quantity
    # Strip leading sl.no if present: "1. " / "1) " — but NOT "1.5" or "1.1/2" (period followed by digit)
    desc = re.sub(r'^\d+[\.\)](?!\d)\s*', '', line).strip()
    if _looks_like_gasket(desc):
        return {'line_no': None, 'description': desc, 'quantity': None, 'uom': 'NOS'}

    return None


def _looks_like_gasket(text: str) -> bool:
    text_lower = text.lower()
    # Must contain gasket-related keyword or size indicator
    gasket_kws = ['gasket', 'gkt', 'rubber', 'ptfe', 'neoprene', 'epdm', 'cnaf',
                  'viton', 'graphite', 'grph', 'graph fill', 'pn', '150#', '300#', '600#',
                  'asme', 'ansi', 'b16.20', 'b16.21',
                  'rtj', 'r.t.j', 'ring joint', 'joint tore', 'tore', 'spiral', 'winding',
                  'spw', 'wnd', 'sw gasket', 'kammprofile', 'camprofile', 'insulating gasket',
                  'isk', 'soft iron', 'softiron', 'octagonal', 'oval ring',
                  'nbr', 'nitrile', 'sbr', 'silicone', 'butyl', 'aramid', 'thermiculite',

                  'expanded graphite', 'cork', 'leather', 'ceramic fiber', 'hnbr',
                  'outer ring', 'inner ring', 'centering ring', 'gid', 'god']
    has_kw = any(k in text_lower for k in gasket_kws)
    has_size = bool(re.search(
        r'\d+["\']|\d+\s*(?:nb|dn|nps|inch|mm)|(?:nb|dn)\s*\d+|\d+\s*(?:gid|god)\b',
        text_lower, re.IGNORECASE))
    return has_kw or has_size


def _normalize_uom(uom: str) -> str:
    uom = uom.strip().upper()
    if uom.startswith('M') and not uom.startswith('MT'):
        return 'M'
    return 'NOS'


_AI_LAYOUT_PROMPT = """\
You are analysing a customer Excel sheet that contains gasket procurement enquiries.

Below are the first rows of one worksheet (each object has "row" = 1-based row index and "cells" = list of cell values left-to-right, empty string for blank cells).

{preview_json}

Your task: identify the layout so downstream code can extract every gasket line item.

Return ONLY valid JSON (no markdown, no explanation) matching this schema exactly:
{{
  "header_row": <1-based integer row that contains column headers, or null if no header>,
  "format_type": "description_column" | "structured_columns",
  "columns": {{
    "description":    <0-based column index of the full-text gasket description, or null>,
    "component_name": <0-based column index whose DATA values indicate the gasket TYPE (e.g. "Gasket SPW", "Spiral Wound Gasket", "RTJ Ring", "Soft Cut Gasket") — short type-label cells, not full specs; or null>,
    "size":           <0-based column index for nominal size (NB/DN/NPS/inch), or null>,
    "rating":         <0-based column index for pressure rating/class, or null>,
    "moc":            <0-based column index for material / MOC, or null>,
    "thickness":      <0-based column index for thickness (mm), or null>,
    "od_mm":          <0-based column index for OD in mm, or null>,
    "id_mm":          <0-based column index for ID in mm, or null>,
    "face_type":      <0-based column index for face type (RF/FF), or null>,
    "quantity":       <0-based column index for quantity / qty, or null>,
    "uom":            <0-based column index for unit of measure, or null>,
    "line_no":        <0-based column index for serial / item number, or null>
  }}
}}

IMPORTANT — choose format_type by looking at the ACTUAL DATA VALUES, not just column headers:

"description_column" — use this when a column's DATA contains long, complete gasket
specifications (full phrases with size + rating + material + type together), e.g.:
  "6 inch 150# CNAF RF Gasket"
  "Spiral wound graphite SS316 gasket 10\\" 300#"
  "8in, Gasket Spiral Wound, 4.5mm Thk, CL150, WD-SS316, IR-Graphite, OR-CS, ASME B16.20"
Map that column as "description". Also map quantity, uom, line_no from their columns.
Even if other columns (size, moc, qty) also exist, prefer "description_column" when one
column already contains the complete specification.

"structured_columns" — use this ONLY when NO column contains the full spec in one cell,
and separate columns each carry one field (size in one col, rating in another, moc in another).

Key rules:
- Look at the actual cell values in data rows to decide: long comma-separated text → description_column.
- Short type-label values (like "Gasket SPW", "Spiral Wound", "RTJ") → map as "component_name"
  (not "description") so the gasket type is preserved even in structured_columns mode.
- Set a column index only when you are confident it maps to that field; null otherwise.
- The header_row may be in rows 1–15; data rows start immediately after.
"""


def _ai_detect_sheet_layout(ws, openai_client) -> dict | None:
    """Send the first 20 rows to the LLM and return the column layout dict, or None on failure."""
    preview_rows = []
    for row_idx, row in enumerate(worksheet_rows_with_merged_values(ws, max_row=20), 1):
        cells = [str(c) if c is not None else '' for c in row]
        if any(s.strip() for s in cells):
            preview_rows.append({'row': row_idx, 'cells': cells})

    if not preview_rows:
        return None

    prompt = _AI_LAYOUT_PROMPT.format(preview_json=json.dumps(preview_rows, ensure_ascii=False))
    try:
        response = openai_client.chat.completions.create(
            model='gpt-4o-mini',
            temperature=0,
            response_format={'type': 'json_object'},
            messages=[{'role': 'user', 'content': prompt}],
            timeout=30,
        )
        raw = response.choices[0].message.content or ''
        layout = json.loads(raw)
        # Basic sanity: must have 'columns' dict
        if not isinstance(layout.get('columns'), dict):
            return None
        logger.info('AI sheet layout detected: %s', layout)
        return layout
    except Exception as exc:
        logger.warning('AI sheet layout detection failed: %s', exc)
        return None


def _extract_items_from_ai_layout(ws, layout: dict) -> list[dict]:
    """Given an AI-detected layout, extract all gasket line items from the worksheet."""
    header_row = layout.get('header_row') or 1
    fmt = layout.get('format_type', 'description_column')
    cols = layout.get('columns', {})

    def ci(field):
        """Return 0-based column index for a field, or None."""
        v = cols.get(field)
        return int(v) if v is not None else None

    items = []

    if fmt == 'description_column':
        desc_col = ci('description')
        if desc_col is None:
            return []
        qty_col = ci('quantity')
        uom_col = ci('uom')
        line_col = ci('line_no')
        moc_col = ci('moc')

        rows = worksheet_rows_with_merged_values(ws)
        for row in rows[header_row:]:
            desc = _cell_str(row, desc_col)
            if not desc or not _looks_like_gasket(desc):
                continue
            if moc_col is not None:
                moc_val = _cell_str(row, moc_col)
                if moc_val and moc_val.upper() != desc.upper():
                    desc = desc + ' MOC: ' + moc_val
            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            items.append({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'quantity': qty,
                'uom': uom,
            })

    else:  # structured_columns
        size_col = ci('size')
        rating_col = ci('rating')
        moc_col = ci('moc')
        thk_col = ci('thickness')
        od_col = ci('od_mm')
        id_col = ci('id_mm')
        face_col = ci('face_type')
        qty_col = ci('quantity')
        uom_col = ci('uom')
        line_col = ci('line_no')
        comp_col = ci('component_name')

        last_moc = None
        last_comp = None

        rows = worksheet_rows_with_merged_values(ws)
        for row in rows[header_row:]:
            # Fill-down MOC and component_name (often repeated only in first row)
            mat_val = _cell_str(row, moc_col) if moc_col is not None else None
            if mat_val:
                last_moc = mat_val
            else:
                mat_val = last_moc

            comp_val = _cell_str(row, comp_col) if comp_col is not None else None
            if comp_val:
                last_comp = comp_val
            else:
                comp_val = last_comp

            # Build synthetic description
            desc_parts = []

            # Prepend component type label (e.g. "Gasket SPW") so gasket type
            # is preserved for the extractor — this is the most important field
            if comp_val:
                desc_parts.append(comp_val)

            if od_col is not None and id_col is not None:
                od = _cell_str(row, od_col)
                id_ = _cell_str(row, id_col)
                if od and id_:
                    thk = _cell_str(row, thk_col) if thk_col is not None else None
                    thk_part = f' X {thk}MM THK' if thk else ''
                    desc_parts.append(f'OD {od}MM X ID {id_}MM{thk_part}')
            elif size_col is not None:
                size_raw = _cell_str(row, size_col)
                if size_raw:
                    try:
                        sv = float(size_raw)
                        size_str = f'{int(sv)}"' if sv == int(sv) else f'{sv}"'
                    except ValueError:
                        size_str = size_raw
                    rating_raw = _cell_str(row, rating_col) if rating_col is not None else None
                    thk = _cell_str(row, thk_col) if thk_col is not None else None
                    thk_part = f' {thk}MM THK' if thk else ''
                    rating_part = f' {rating_raw}' if rating_raw else ''
                    desc_parts.append(f'{size_str}{rating_part}{thk_part}')

            # Need at least a size/dimension — comp_val alone is not enough
            has_dimension = len(desc_parts) > (1 if comp_val else 0)
            if not has_dimension:
                continue

            if mat_val:
                desc_parts.append(f'MOC: {mat_val}')
            if face_col is not None:
                face_val = _cell_str(row, face_col)
                if face_val:
                    desc_parts.append(face_val)

            desc = ' '.join(desc_parts)
            # Append GASKET suffix only if no gasket-type keyword already present
            if 'gasket' not in desc.lower():
                desc += ' GASKET'
            if not _looks_like_gasket(desc):
                continue

            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            items.append({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'quantity': qty,
                'uom': uom,
            })

    return items


def parse_excel_file(file_bytes: bytes, openai_client=None) -> list[dict]:
    """Parse an uploaded Excel file into raw line items.

    If openai_client is provided, uses AI to detect column layout for each sheet
    (robust against any column naming). Falls back to rule-based detection otherwise.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_items = _parse_sheet(ws, openai_client=openai_client)
        if sheet_items:
            items.extend(sheet_items)
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
    return items


def excel_requires_smart_parse(file_bytes: bytes) -> bool:
    """Return True for layouts where the app should not use Excel fast path."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if any(rng.min_row != rng.max_row for rng in ws.merged_cells.ranges):
            return True

        rows = worksheet_rows_with_merged_values(ws)
        description_sections = _detect_description_sections(rows)
        structured_sections = _detect_structured_sections(rows)
        if len(description_sections) + len(structured_sections) > 1:
            return True

    return False


def _parse_sheet(ws, openai_client=None) -> list[dict]:
    """Detect header row and extract line items from a worksheet.

    With openai_client: uses AI to identify column layout — works for any format.
    Without openai_client: falls back to keyword-based detection.
    """
    # --- AI-powered layout detection (preferred) ---
    if openai_client is not None:
        layout = _ai_detect_sheet_layout(ws, openai_client)
        if layout:
            try:
                items = _extract_items_from_ai_layout(ws, layout)
            except Exception as exc:
                logger.warning('AI item extraction failed, falling back to rule-based: %s', exc)
                items = []
            if items:
                return items
            # AI detected a layout but found no items — fall through to rule-based

    # --- Rule-based fallback: description-based format ---
    items = _parse_description_sections(ws)
    if items:
        return items

    # --- Rule-based fallback: structured column format ---
    return _parse_structured_sheet(ws)


def _norm_header_cell(cell) -> str:
    """Normalize a header cell: lowercase, collapse whitespace incl. non-breaking spaces."""
    if cell is None:
        return ''
    return re.sub(r'[\xa0\s]+', ' ', str(cell)).strip().lower()


def _classify_structured_col(norm: str) -> str | None:
    """Return the column type for a normalized header cell, or None if unrecognised."""
    if norm in ('dn', 'nb'):
        return 'dn_size'
    if norm == 'class':
        return 'class_rating'
    if norm in ('material', 'moc'):
        return 'material'
    if norm in ('thickness', 'thk', 'thk (mm)', 'thick'):
        return 'thickness'
    if norm.startswith('(od)') or norm.startswith('od (') or norm.startswith('od mm') or norm == 'od':
        return 'od_mm'
    if norm.startswith('(id)') or norm.startswith('id (') or norm.startswith('id mm') or norm == 'id':
        return 'id_mm'
    # SIZE column (inches): "SIZE", "SIZE (INCH)", "SIZE\n(INCH)", "SIZE (MM)", "NPS"
    if norm in ('size', 'nps') or norm.startswith('size (') or norm.startswith('size\n'):
        return 'size_inch'
    # RATING column: "RATING", "RATING ", "PRESSURE RATING", "CLASS/RATING"
    if norm in ('rating', 'pressure rating', 'class/rating', 'rating (class)') or norm.startswith('rating'):
        return 'rating'
    if norm in ('qty', 'quantity'):
        return 'quantity'
    if norm in ('uom', 'inv uom'):
        return 'uom'
    if 'sl.no' in norm or 'sr.no' in norm or norm == 'sno' or norm == 'sl no' or norm == 'sr no':
        return 'line_no'
    return None


def _classify_description_col(norm: str) -> str | None:
    if not norm:
        return None
    for col_type, keywords in _HEADER_PATTERNS.items():
        if any(kw in norm for kw in keywords):
            return col_type
    if norm == 'material':
        return 'moc'
    return _classify_structured_col(norm)


def _detect_description_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all description-column header blocks in a sheet."""
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_description_header(col_map: dict) -> bool:
        return 'description' in col_map and (
            'quantity' in col_map or 'moc' in col_map or 'size_inch' in col_map
            or 'rating' in col_map or ('od_mm' in col_map and 'id_mm' in col_map)
        )

    for row_idx, row in enumerate(all_rows):
        col_map = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            col_type = _classify_description_col(norm)
            if col_type and col_type not in col_map:
                col_map[col_type] = col_idx

        if _is_description_header(col_map):
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None and any(c is not None for c in row):
            current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _append_field(parts: list[str], label: str, value: str | None, suffix: str = '') -> None:
    if value:
        if re.fullmatch(r'n/?a|not applicable|nil|none', value.strip(), re.IGNORECASE):
            return
        parts.append(f'{label}: {value}{suffix}')


def _description_from_section_row(row: tuple, col_map: dict) -> str | None:
    desc = _cell_str(row, col_map.get('description'))
    if not desc or not _looks_like_gasket(desc):
        return None

    parts = [desc]
    size = _cell_str(row, col_map.get('size_inch'))
    rating = _cell_str(row, col_map.get('rating'))
    thk = _cell_str(row, col_map.get('thickness'))
    od = _cell_str(row, col_map.get('od_mm'))
    id_ = _cell_str(row, col_map.get('id_mm'))

    if size:
        size_rating = size
        if rating:
            size_rating += f' X {rating}'
        if thk:
            size_rating += f' X {thk}MM THK'
        parts.append(size_rating)
    elif od and id_:
        dim = f'OD {od}MM X ID {id_}MM'
        if thk:
            dim += f' X {thk}MM THK'
        parts.append(dim)
    elif thk:
        _append_field(parts, 'THK', thk, 'MM')

    moc_val = _cell_str(row, col_map.get('moc'))
    if moc_val and moc_val.upper() != desc.upper():
        parts.append(f'MOC: {moc_val}')

    _append_field(parts, 'INNER RING WIDTH', _cell_str(row, col_map.get('inner_ring_width')), 'MM')
    _append_field(parts, 'OUTER RING WIDTH', _cell_str(row, col_map.get('outer_ring_width')), 'MM')
    _append_field(parts, 'REMARKS', _cell_str(row, col_map.get('remarks')))

    return ' '.join(parts)


def _parse_description_sections(ws) -> list[dict]:
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_description_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')
        if line_col is None:
            line_col = _infer_line_no_col(data_rows, col_map.get('description'))

        for row in data_rows:
            desc = _description_from_section_row(row, col_map)
            if not desc:
                continue
            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            if qty_col is not None and qty is None and line_no is None:
                continue
            items.append({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'quantity': qty,
                'uom': uom,
            })

    return items


def _infer_line_no_col(data_rows: list[tuple], desc_col: int | None) -> int | None:
    """Infer a serial-number column immediately to the left of Description."""
    if desc_col is None or desc_col <= 0:
        return None
    best_col = None
    best_count = 0
    for col_idx in range(desc_col):
        count = 0
        for row in data_rows[:25]:
            if _cell_float(row, col_idx) is not None:
                count += 1
        if count > best_count:
            best_col = col_idx
            best_count = count
    return best_col if best_count >= 1 else None


def _detect_structured_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all structured-format header blocks in a sheet.

    Returns a list of (header_row_idx_0based, col_map, data_rows) tuples where
    col_map maps field names to 0-based column indices and data_rows are the
    value tuples between this header and the next.
    """
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_structured_header(col_map: dict) -> bool:
        has_material = 'material' in col_map
        has_dn_cls = 'dn_size' in col_map and 'class_rating' in col_map
        has_od_id = 'od_mm' in col_map and 'id_mm' in col_map
        has_size_rating = 'size_inch' in col_map and 'rating' in col_map
        return has_material and (has_dn_cls or has_od_id or has_size_rating)

    for row_idx, row in enumerate(all_rows):
        # Try to interpret this row as a header
        col_map = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            if not norm:
                continue
            col_type = _classify_structured_col(norm)
            if col_type and col_type not in col_map:
                col_map[col_type] = col_idx

        if _is_structured_header(col_map):
            # Save previous section if any
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None:
            # Accumulate data rows (skip fully-empty rows)
            if any(c is not None for c in row):
                current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _parse_structured_sheet(ws) -> list[dict]:
    """Parse a sheet that uses one column per field (no combined description column)."""
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_structured_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        mat_col = col_map.get('material')
        dn_col = col_map.get('dn_size')
        cls_col = col_map.get('class_rating')
        thk_col = col_map.get('thickness')
        od_col = col_map.get('od_mm')
        id_col = col_map.get('id_mm')
        size_col = col_map.get('size_inch')
        rating_col = col_map.get('rating')
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')

        last_material = None

        for row in data_rows:
            # Fill-down for MATERIAL column
            mat_val = _cell_str(row, mat_col) if mat_col is not None else None
            if mat_val:
                last_material = mat_val
            else:
                mat_val = last_material

            if not mat_val:
                continue

            # Build synthetic description
            if dn_col is not None and cls_col is not None:
                dn = _cell_str(row, dn_col)
                cls = _cell_str(row, cls_col)
                if not dn or not cls:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{dn} NB {cls}#{thk_part} GASKET MOC: {mat_val}'
            elif od_col is not None and id_col is not None:
                od = _cell_str(row, od_col)
                id_ = _cell_str(row, id_col)
                if not od or not id_:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' X {thk}MM THK' if thk else ''
                desc = f'OD {od}MM X ID {id_}MM{thk_part} GASKET MOC: {mat_val}'
            elif size_col is not None and rating_col is not None:
                size_raw = _cell_str(row, size_col)
                rating_raw = _cell_str(row, rating_col)
                if not size_raw or not rating_raw:
                    continue
                # Numeric inch value (e.g. 10 → 10", 1.5 → 1.5")
                try:
                    sv = float(size_raw)
                    size_str = f'{int(sv)}"' if sv == int(sv) else f'{sv}"'
                except ValueError:
                    size_str = size_raw  # already a formatted string
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{size_str} {rating_raw}{thk_part} GASKET MOC: {mat_val}'
            else:
                continue

            if not _looks_like_gasket(desc):
                continue

            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            items.append({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'quantity': qty,
                'uom': uom,
            })

    return items


# Keywords to identify column types — order matters: more specific first
_HEADER_PATTERNS = {
    'description': ['description', 'notes'],
    'quantity':    ['qty', 'quantity', 'balance to order', 'balance', 'required qty', 'count'],
    'uom':         ['uom', 'inv uom'],
    'line_no':     ['sl.no', 'sl no', 'sr. no', 'sr no', 'sr no.', 'sno', 'serial', 'sr. no.', 'sr no'],
    'moc':         ['moc'],
    'size_inch':   ['nps', 'size'],
    'rating':      ['class /rating', 'class/rating', 'rating', 'class'],
    'thickness':   ['thickness', 'thk', 'thick'],
    'od_mm':       ['gasket od', '(od)', 'od (', 'od mm'],
    'id_mm':       ['gasket id', '(id)', 'id (', 'id mm'],
    'remarks':     ['remarks', 'remark'],
    'inner_ring_width': ['inner ring width'],
    'outer_ring_width': ['outer / center ring  width', 'outer / center ring width', 'outer ring width',
                         'center ring width', 'centering ring width'],
}


def _detect_header(ws) -> tuple[int, dict]:
    """Scan first 15 rows to find the header row and map column indices."""
    best_row, best_col_map, best_score = 0, {}, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        col_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_lower = str(cell).lower().strip()
            for col_type, keywords in _HEADER_PATTERNS.items():
                if col_type not in col_map and any(kw in cell_lower for kw in keywords):
                    col_map[col_type] = col_idx  # 0-based
        score = len(col_map)
        if score >= 2 and 'description' in col_map and score > best_score:
            best_score, best_row, best_col_map = score, row_idx, col_map
    return best_row, best_col_map


def _cell_str(row: tuple, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    # Normalize Shift+Enter newlines within a cell into a single space
    return re.sub(r'[\r\n]+', ' ', str(val)).strip()


def _cell_float(row: tuple, idx: int | None) -> float | None:
    if idx is None or idx >= len(row):
        return None
    try:
        return float(row[idx])
    except (TypeError, ValueError):
        return None
