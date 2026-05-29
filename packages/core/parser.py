from __future__ import annotations
"""
Parses customer enquiry inputs (email text or Excel) into a uniform list of raw items:
[{'line_no': int, 'description': str, 'quantity': float, 'uom': str}]
"""
import re
import io
import logging
import openpyxl

logger = logging.getLogger(__name__)


def worksheet_rows_with_merged_values(ws, max_row: int | None = None) -> list[tuple]:
    """Return worksheet rows with merged cells expanded to every covered row.

    openpyxl exposes only the top-left value of a merged range. Enquiry sheets
    often merge MOC/spec cells vertically, so reading values directly makes the
    following line items look incomplete.
    """
    row_limit = min(max_row or ws.max_row, ws.max_row)
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, row_limit + 1)
    ]

    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            continue
        for r in range(merged_range.min_row, min(merged_range.max_row, row_limit) + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                rows[r - 1][c - 1] = value

    return [tuple(row) for row in rows]


def parse_email_text(text: str) -> list[dict]:
    """Extract line items from pasted email body text via regex/rule-based parsing."""
    items = []
    lines = _merge_continuation_lines([l.strip() for l in text.splitlines() if l.strip()])
    for line in lines:
        item = _parse_line(line)
        if item:
            items.append(item)
    # Number items if line_no missing
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
    return items


def _merge_continuation_lines(lines: list[str]) -> list[str]:
    """Merge lines that are mid-sentence continuations of the previous line.

    Handles cases where a long gasket description wraps across two lines,
    e.g. Excel cell content copied as text where a description ends with
    'OUTER RING' and the next line starts with 'MATERIAL: Carbon Steel...'.
    """
    merged = []
    for line in lines:
        if not merged:
            merged.append(line)
            continue
        prev = merged[-1]
        # A line is a continuation if:
        # 1. It does NOT start with a digit (new serial number), AND
        # 2. The previous line ended with a comma, OR ended with RING/MATERIAL/OUTER/INNER
        #    (cut mid-phrase), OR the current line starts with a field-label ("WORD:") pattern
        starts_with_number = bool(re.match(r'^\d', line))
        # Lines starting with a gasket size-prefix keyword are always new items
        is_new_item_prefix = bool(re.match(r'^(?:NPS|NB|DN|SIZE)\s*:', line, re.IGNORECASE))
        prev_ends_mid = (
            prev.endswith(',')
            or re.search(r'\b(?:OUTER\s*RING|INNER\s*RING|OUTER|RING|MATERIAL)\s*$', prev, re.IGNORECASE)
        )
        # Previous line ends with a standards-body acronym — next line is its number (e.g. "ANSI\n B16.47")
        prev_ends_std_prefix = bool(
            re.search(r'\b(?:ANSI|ASME|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\s*$', prev, re.IGNORECASE)
        )
        curr_is_field_continuation = bool(re.match(r'^[A-Z][A-Z\s]+:\s*\S', line))
        # Current line is a standard reference line (e.g. "ASME B16.20: Metallic Gaskets for Pipe Flanges")
        # that belongs to the preceding description, not a new line item
        curr_is_standard_ref = bool(re.match(
            r'^(?:ASME|ANSI|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\b', line, re.IGNORECASE
        ))
        if (not starts_with_number and not is_new_item_prefix
                and (prev_ends_mid or prev_ends_std_prefix or curr_is_field_continuation or curr_is_standard_ref)):
            merged[-1] = prev + ' ' + line
        else:
            merged.append(line)
    return merged


def _parse_line(line: str) -> "dict | None":
    """Try to extract (description, qty, uom) from a single text line."""
    # Skip header-like lines
    lower = line.lower()
    if any(kw in lower for kw in ['sl.no', 'sl no', 'line no', 'release no', 'notes', 'inv uom',
                                   'description', 'subject', 'dear', 'regards', 'kindly',
                                   'total', 'terms', 'price', 'revision', 'date']):
        return None

    # Pattern: optional number, description text, number, unit
    # e.g. "1  Gasket - Rubber - 6'' PN10  27  m"
    pattern = re.compile(
        r'^(\d+)?\s*'                            # optional sl.no
        r'(?:\d+\s+\d+\s+)?'                     # optional line/release no columns
        r'(.+?)\s+'                              # description (greedy)
        r'(\d+(?:\.\d+)?)\s*'                    # quantity
        r'(nos?|m|mtr|meters?|pcs?|sets?|kgs?|units?)\s*$',  # unit
        re.IGNORECASE
    )
    m = pattern.match(line)
    if m:
        sl_no = int(m.group(1)) if m.group(1) else None
        desc = m.group(2).strip()
        qty = float(m.group(3))
        uom = _normalize_uom(m.group(4))
        if _looks_like_gasket(desc) and qty > 0:
            return {'line_no': sl_no, 'description': desc, 'quantity': qty, 'uom': uom}

    # Simpler fallback: tab/multiple-space separated columns
    parts = re.split(r'\t|  +', line)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 3:
        # Try last two parts as qty + uom, rest as description
        try:
            qty = float(parts[-2])
            uom = _normalize_uom(parts[-1])
            desc = ' '.join(parts[:-2]).strip()
            desc = re.sub(r'^\d+\s+', '', desc).strip()
            if _looks_like_gasket(desc) and qty > 0:
                return {'line_no': None, 'description': desc, 'quantity': qty, 'uom': uom}
        except (ValueError, IndexError):
            pass

    # Last resort: entire line is a description with no quantity
    # Strip leading sl.no if present: "1. " / "1) " — but NOT "1.5" or "1.1/2" (period followed by digit)
    desc = re.sub(r'^\d+[\.\)](?!\d)\s*', '', line).strip()
    if _looks_like_gasket(desc):
        return {'line_no': None, 'description': desc, 'quantity': None, 'uom': 'NOS'}

    return None


def _looks_like_gasket(text: str) -> bool:
    text_lower = text.lower()
    # Must contain gasket-related keyword or size indicator
    gasket_kws = ['gasket', 'gkt', 'rubber', 'ptfe', 'neoprene', 'epdm', 'cnaf',
                  'viton', 'graphite', 'grph', 'graph fill', 'pn', '150#', '300#', '600#',
                  'asme', 'ansi', 'b16.20', 'b16.21',
                  'rtj', 'r.t.j', 'ring joint', 'joint tore', 'tore', 'spiral', 'winding',
                  'spw', 'wnd', 'sw gasket', 'kammprofile', 'kamprofile', 'camprofile', 'insulating gasket',
                  'insulation gasket', 'isolating kit', 'insulating kit',
                  'isk', 'soft iron', 'softiron', 'octagonal', 'oval ring',
                  'nbr', 'nitrile', 'sbr', 'silicone', 'butyl', 'aramid', 'thermiculite',

                  'expanded graphite', 'cork', 'leather', 'ceramic fiber', 'hnbr',
                  'outer ring', 'inner ring', 'centering ring', 'gid', 'god']
    has_kw = any(k in text_lower for k in gasket_kws)
    has_size = bool(re.search(
        r'\d+["\']|\d+\s*(?:nb|dn|nps|inch|mm)|(?:nb|dn)\s*\d+|\d+\s*(?:gid|god)\b',
        text_lower, re.IGNORECASE))
    return has_kw or has_size


def _normalize_uom(uom: str) -> str:
    uom = uom.strip().upper()
    if uom.startswith('M') and not uom.startswith('MT'):
        return 'M'
    return 'NOS'


def parse_excel_file(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded Excel file into raw line items via rule-based detection."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_items = _parse_sheet(ws)
        if sheet_items:
            items.extend(sheet_items)
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
        item.setdefault('source_index', i)
    return items


def excel_requires_smart_parse(file_bytes: bytes) -> bool:
    """Return True for layouts where the app should not use Excel fast path."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if any(rng.min_row != rng.max_row for rng in ws.merged_cells.ranges):
            return True

        rows = worksheet_rows_with_merged_values(ws)
        description_sections = _detect_description_sections(rows)
        structured_sections = _detect_structured_sections(rows)
        if len(description_sections) + len(structured_sections) > 1:
            return True

    return False


def _parse_sheet(ws) -> list[dict]:
    """Detect header row and extract line items from a worksheet (rule-based)."""
    items = _parse_description_sections(ws)
    if items:
        return items
    return _parse_structured_sheet(ws)


def _norm_header_cell(cell) -> str:
    """Normalize a header cell: lowercase, collapse whitespace incl. non-breaking spaces."""
    if cell is None:
        return ''
    return re.sub(r'[\xa0\s]+', ' ', str(cell)).strip().lower()


def _classify_structured_col(norm: str) -> str | None:
    """Return the column type for a normalized header cell, or None if unrecognised."""
    if norm in ('dn', 'nb'):
        return 'dn_size'
    if norm == 'class':
        return 'class_rating'
    if norm in ('material', 'moc'):
        return 'material'
    if norm in ('thickness', 'thk', 'thk (mm)', 'thick'):
        return 'thickness'
    if norm.startswith('(od)') or norm.startswith('od (') or norm.startswith('od mm') or norm == 'od':
        return 'od_mm'
    if norm.startswith('(id)') or norm.startswith('id (') or norm.startswith('id mm') or norm == 'id':
        return 'id_mm'
    # SIZE column (inches): "SIZE", "SIZE (INCH)", "SIZE\n(INCH)", "SIZE (MM)", "NPS"
    if norm in ('size', 'nps') or norm.startswith('size (') or norm.startswith('size\n'):
        return 'size_inch'
    # RATING column: "RATING", "RATING ", "PRESSURE RATING", "CLASS/RATING"
    if norm in ('rating', 'pressure rating', 'class/rating', 'rating (class)') or norm.startswith('rating'):
        return 'rating'
    if norm in ('qty', 'quantity'):
        return 'quantity'
    if norm in ('uom', 'inv uom'):
        return 'uom'
    if 'sl.no' in norm or 'sr.no' in norm or norm == 'sno' or norm == 'sl no' or norm == 'sr no':
        return 'line_no'
    return None


def _classify_description_col(norm: str) -> str | None:
    if not norm:
        return None
    if norm in ('piping class', 'pipe class'):
        return None
    for col_type, keywords in _HEADER_PATTERNS.items():
        if any(kw in norm for kw in keywords):
            return col_type
    if norm == 'material':
        return 'moc'
    return _classify_structured_col(norm)


def _detect_description_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all description-column header blocks in a sheet."""
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_description_header(col_map: dict) -> bool:
        return 'description' in col_map and (
            'quantity' in col_map or 'moc' in col_map or 'size_inch' in col_map
            or 'rating' in col_map or ('od_mm' in col_map and 'id_mm' in col_map)
        )

    for row_idx, row in enumerate(all_rows):
        col_map = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            col_type = _classify_description_col(norm)
            if col_type and col_type not in col_map:
                col_map[col_type] = col_idx

        if _is_description_header(col_map):
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None and any(c is not None for c in row):
            current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _append_field(parts: list[str], label: str, value: str | None, suffix: str = '') -> None:
    if value:
        if re.fullmatch(r'n/?a|not applicable|nil|none', value.strip(), re.IGNORECASE):
            return
        parts.append(f'{label}: {value}{suffix}')


def _float_from_text(value: str | None) -> float | None:
    if not value:
        return None
    match = re.search(r'\d+(?:\.\d+)?', str(value))
    return float(match.group(0)) if match else None


def _rating_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip().upper()
    raw = re.sub(r'(PN)\s*(\d+)', r'\1 \2', raw)
    pn = re.search(r'\bPN\s*-?\s*(\d{1,2})(?!\d)', raw)
    if pn:
        return f'PN {pn.group(1)}'
    lb = re.search(r'\b(150|300|400|600|900|1500|2500)\s*LB\b', raw)
    if lb:
        return f'{lb.group(1)}#'
    dual_cls = re.search(
        r'\b(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\s*'
        r'(?:/|\\|OR|AND|-)\s*'
        r'(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\b',
        raw,
    )
    if dual_cls:
        return f'{dual_cls.group(1)}/{dual_cls.group(2)}#'
    cls = re.search(r'\b(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\b', raw)
    if cls:
        return f'{cls.group(1)}#'
    return raw if len(raw) <= 20 else None


def _size_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    fraction_map = {
        '¼': '0.25',
        '½': '0.5',
        '¾': '0.75',
    }
    if raw in fraction_map:
        raw = fraction_map[raw]
    try:
        number = float(raw)
        return f'{int(number)}"' if number == int(number) else f'{number}"'
    except ValueError:
        return raw


def _infer_gasket_type(description: str) -> str | None:
    raw = description.upper()
    raw = re.sub(r'(?i)(GASKET)(?=(SKAG|CAM|KAMM|DOUBLE|COPPER|\d))', r'\1 ', raw)
    raw = re.sub(r'(?i)(\d)(INST\.?\s+KIT|INSULATING|IN\s+GASKET)', r'\1 \2', raw)
    if re.search(
        r'INSULAT(?:ING|ION)|\bNSULATING\b|\bISK\b|\bFLANGE\s+ISOLATION\b|'
        r'\bFLANGE\s+INSULATION\s+KIT\b|\bFLANGE\s+ISOLAT(?:ING|ION)\s+KIT\b|'
        r'\bINSULATION\s+KIT\b|\bISOLAT(?:ING|ION)\s+GASKET\b|\bCOMMANDER\s+EXTREME\b|\bINST\.?\s+KIT\b',
        raw,
    ):
        if re.search(r'\bRTJ\b|\bR/?J\b|\bRING\s+JOINT\b|\bTYPE[-\s]?D\b', raw):
            return 'ISK_RTJ'
        return 'ISK'
    if re.search(
        r'\b(?:SPIRAL|SPRIAL|SPRIRAL|SPIRIAL|SPLRAL|SPRLAL|SPIRRAL|SPRRAL|SPRL)\s*[-\s]*(?:W(?:OU)?ND\w*|WIND\w*)\b'
        r'|\bSPW(?:D)?\b|\bSW\s+GASKET\b',
        raw,
    ):
        return 'SPIRAL_WOUND'
    if re.search(r'\b(?:RING\s+JOINT|RING\s+TYPE\s+JOINT|RING\s+TYPE\s+GASKET|RTJ|RJ\s+GASKET|R/?J)\b|'
                 r'\b(?:OCTAGONAL|OVAL)\s+RING\s+GASKETS?\b', raw):
        return 'RTJ'
    if re.search(r'\bKAMMPROFILE\b|\bKAMPROFILE\b|\bKAMM\s*PROFILE\b|\bCAMPROFILE\b|\bCAM\s*PROFILE\b|\bPROFILE\s+GASKET\b|\bGROOVED\s+PROFILE\b|\bGROOVED\s+METAL\b|SKAG', raw):
        return 'KAMM'
    if re.search(r'\bDOUBLE[\s\-]?JACKET(?:ED)?\b|\bJACKETED\b|\bJACKET\s+GASKET\b|\bCOPPER\s+JACKET\b', raw):
        return 'DJI'
    if re.search(r'\bPLUG\s+GASKET\b|\bPLUG\s+TYPE\s+GASKET\b', raw):
        return 'PLUG_GASKET'
    if re.search(r'\bCORRUGATED(?:\s+METAL(?:LIC)?)?\s+GASKET\b|\bCORRUGATED\s+GASKET\b', raw):
        return 'CORRUGATED'
    if re.search(r'\bSHEET\s+GASKET\b|\bGASKET\s+SHEET\b', raw):
        return 'SHEET_GASKET'
    if re.search(r'\bGASKET\b.*\bO\.?D\.?\s*\d+.*\bI\.?D\.?\s*\d+', raw, re.IGNORECASE) and re.search(r'\bDRAWING\b|\bPOSITION\b|\bASBESTOS\s+FREE\b', raw):
        return 'DJI'
    if _looks_like_gasket(description):
        return 'SOFT_CUT'
    return None


def _standard_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).upper()
    match = re.search(r'\b(?:ASME|ANSI)\s+B\s*16\.(20|21|47)\b', raw)
    if match:
        return f'ASME B16.{match.group(1)}'
    match = re.search(r'\bAPI\s*6A\b', raw)
    if match:
        return 'API 6A'
    match = re.search(r'\bEN\s*1514[-\s]*(\d+)\b', raw)
    if match:
        return f'EN 1514-{match.group(1)}'
    return None


_SW_MATERIAL_ALIASES: list[tuple[str, str]] = [
    (r'SS\s*TP\s*316\s*/\s*316L', 'SS316/316L'),
    (r'TP\s*316\s*/\s*316L', 'SS316/316L'),
    (r'SS\s*316\s*/\s*SS\s*316L', 'SS316/316L'),
    (r'SS\s*316\s*/\s*316L', 'SS316/316L'),
    (r'S\.?\s*S\.?\s*316L', 'SS316L'),
    (r'S\.?\s*S\.?\s*316', 'SS316'),
    (r'S\.?\s*S\.?\s*304L', 'SS304L'),
    (r'S\.?\s*S\.?\s*304', 'SS304'),
    (r'TP\s*316L\s*SS', 'SS316L'),
    (r'TP\s*316\s*SS', 'SS316'),
    (r'TP\s*304L\s*SS', 'SS304L'),
    (r'TP\s*304\s*SS', 'SS304'),
    (r'316L\s*SS', 'SS316L'),
    (r'316\s*SS', 'SS316'),
    (r'304L\s*SS', 'SS304L'),
    (r'304\s*SS', 'SS304'),
    (r'STAINLESS\s+STEEL\s+316L', 'SS316L'),
    (r'STAINLESS\s+STEEL\s+316', 'SS316'),
    (r'STAINLESS\s+STEEL\s+304L', 'SS304L'),
    (r'STAINLESS\s+STEEL\s+304', 'SS304'),
    (r'AISI\s*316L', 'SS316L'),
    (r'AISI\s*316', 'SS316'),
    (r'AISI\s*304L', 'SS304L'),
    (r'AISI\s*304', 'SS304'),
    (r'TP\s*316L', 'SS316L'),
    (r'TP\s*316', 'SS316'),
    (r'TP\s*304L', 'SS304L'),
    (r'TP\s*304', 'SS304'),
    (r'SS\s*317L', 'SS317L'),
    (r'SS\s*317', 'SS317'),
    (r'SS\s*316L', 'SS316L'),
    (r'SS\s*316', 'SS316'),
    (r'SS\s*304L', 'SS304L'),
    (r'SS\s*304', 'SS304'),
    (r'\b316L\b', 'SS316L'),
    (r'\b316\b', 'SS316'),
    (r'\b304L\b', 'SS304L'),
    (r'\b304\b', 'SS304'),
    (r'INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|UNS\s*N08825', 'ALLOY 825'),
    (r'INCONEL\s*625|ALLOY\s*625|UNS\s*N06625', 'INCONEL 625'),
    (r'HASTELLOY\s*C[-\s]*276', 'HASTELLOY C276'),
    (r'DSS\s+UNS\s*S\s*32205|UNS\s*S\s*32205|DUPLEX\s+S32205', 'UNS S32205'),
    (r'DSS\s+UNS\s*S\s*31803|UNS\s*S\s*31803|DUPLEX', 'UNS S31803'),
    (r'SUPER\s+DUPLEX|UNS\s*S\s*32750', 'UNS S32750'),
    (r'UNS\s*S\s*32760', 'UNS S32760'),
    (r'CARBON\s+STEEL|C\.?\s*S\.?|MILD\s+STEEL|M\.?\s*S\.?', 'CS'),
    (r'SOFT\s+IRON', 'SOFT IRON'),
    (r'\bCOPPER\b', 'COPPER'),
    (r'\bLTCS\b', 'LTCS'),
    (r'\bSS\b', 'SS'),
]

_SW_FILLER_ALIASES: list[tuple[str, str]] = [
    (r'EXFOLIATED\s+EXPANDED\s+GRAPHITE', 'EXFOLIATED EXPANDED GRAPHITE'),
    (r'EXPANDED\s+GRAPHITE', 'EXPANDED GRAPHITE'),
    (r'FLEXIBLE\s+GRAPHITE|FLEX\s+GRAPHITE|GRAFOIL|GRAFIL|\bGPH\b|GRAPH(?:ITE|OIL)', 'GRAPHITE'),
    (r'\bPTFE\b|TEFLON', 'PTFE'),
    (r'\bMICA\b', 'MICA'),
    (r'\bCERAMIC\b', 'CERAMIC'),
]

_SW_MATERIAL_RE = '|'.join(f'(?:{pattern})' for pattern, _ in _SW_MATERIAL_ALIASES)
_SW_FILLER_RE = '|'.join(f'(?:{pattern})' for pattern, _ in _SW_FILLER_ALIASES)


def _sw_prepare_text(description: str) -> str:
    text = description.upper()
    text = text.replace('\xa0', ' ')
    typo_replacements = {
        r'\b(?:SPRL|SPRIAL|SPRIRAL|SPIRIAL|SPLRAL|SPRLAL|SPIRRAL|SPRRAL)\s*[-\s]*(?:W(?:OU)?ND\w*|WIND\w*)\b': 'SPIRAL WOUND',
        r'\b(?:WNDLNG|WNDNG|WLDNG|WLNDNG|WINDLNG|WINDNG)\b': 'WINDING',
        r'\b(?:RLNG|RFNG|RNG)\b': 'RING',
        r'\b(?:CENTERLNG|CENTRLNG)\b': 'CENTERING',
        r'\bCENTRING\b': 'CENTERING',
        r'\bSPW(?:D)?\b': 'SPIRAL WOUND',
        r'\bGRAPH(?:L|I)?TLE\b': 'GRAPHITE',
        r'\bFLD\b': 'FILLED',
        r'\bSULT\b': 'SUIT',
    }
    for pattern, replacement in typo_replacements.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = re.sub(r'\bS[GC]\s*(304L?|316L?|317L?)\b', r'SS\1', text, flags=re.IGNORECASE)
    text = re.sub(r'(?i)(GASKET)(?=(SKAG|CAM|KAMM|DOUBLE|COPPER|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(SKAG|KAMMPROFILE|CAMPROFILE|PROFILE)(?=(WITH|FOR|OD|ID|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(THK)(?=(CL|CLASS|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(CL\.?\s*\d{2,4})(?=(FOR|PRO|GASKET|SS|TP|AISI))', r'\1 ', text)
    # Customer exports often lose spaces at boundaries such as FillerAlloy or IRAlloy.
    text = re.sub(r'(?i)(FILLER|FILLED|WINDINGS?|WOUND|CL\d{2,4}|IR|OR)(?=(CL\d{2,4}|ALLOY|INCOLOY|INCONEL|SS|STAINLESS|AISI|TP|CS|CARBON|DSS|UNS))', r'\1 ', text)
    text = re.sub(r'(?i)(GRAPHITE|GRAFOIL|PTFE)(?=(FILLER|FILLED))', r'\1 ', text)
    text = re.sub(r'(?i)(\d)(IN\b)', r'\1 IN', text)
    text = re.sub(r'(?i)(\d)(INST\.?\s+KIT|INSULATING|ISOLATING)', r'\1 \2', text)
    text = re.sub(r'(?i)(EN\s*1514)(\d+)(PN\s*\d+)', r'\1 \2 \3', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _sw_norm_material(raw: str | None) -> str | None:
    if not raw:
        return None
    value = re.sub(r'\s+', ' ', raw.strip().upper())
    value = value.strip(' ,:+-/')
    for pattern, canonical in _SW_MATERIAL_ALIASES:
        if re.fullmatch(pattern, value, re.IGNORECASE):
            return canonical
    for pattern, canonical in _SW_MATERIAL_ALIASES:
        if re.search(r'\b' + pattern + r'\b', value, re.IGNORECASE):
            return canonical
    return value or None


def _sw_norm_filler(raw: str | None) -> str | None:
    if not raw:
        return None
    value = re.sub(r'\s+', ' ', raw.strip().upper())
    for pattern, canonical in _SW_FILLER_ALIASES:
        if re.search(pattern, value, re.IGNORECASE):
            return canonical
    return value or None


def _first_match_material(patterns: list[str], text: str) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return _sw_norm_material(match.group('mat'))
    return None


def _first_match_filler(patterns: list[str], text: str) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return _sw_norm_filler(match.group('mat'))
    return None


def _extract_spw_components(description: str) -> dict:
    """Extract SPW construction from broad customer wording.

    The extractor is intentionally evidence-driven: values are only populated
    when they appear near winding/filler/IR/OR/inner/outer ring cues.
    """
    text = _sw_prepare_text(description)
    result: dict = {}

    size = _extract_first_size(text)
    if size:
        result['size'] = size
        result['size_type'] = 'NPS'

    winding = _first_match_material([
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:WINDINGS?|WOUND\b)',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+SPIRAL\s+WOUND\b',
        rf'(?:WINDINGS?|WOUND|MOC\s*:?)\s+(?P<mat>{_SW_MATERIAL_RE})',
        rf'SPIRAL\s+WOUND\s+(?P<mat>{_SW_MATERIAL_RE})',
    ], text)
    if winding:
        result['sw_winding_material'] = winding

    filler = _first_match_filler([
        rf'(?P<mat>{_SW_FILLER_RE})\s+(?:FILLER|FILLED|FILL\b)',
        rf'(?:FILLER|FILLED|FILL)\s+(?P<mat>{_SW_FILLER_RE})',
    ], text)
    if filler:
        result['sw_filler'] = filler

    inner = _first_match_material([
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:INNER\s+RING|I\.?R\.?\b|IR\b)',
        rf'(?:INNER\s+RING|I\.?R\.?\b|IR\b)\s*[-:]*\s*(?P<mat>{_SW_MATERIAL_RE})',
    ], text)
    outer = _first_match_material([
        rf'(?:OUTER\s+RING|CENTER(?:ING)?\s+RING|CENTRE\s+RING|O\.?R\.?\b|OR\b)\s*[-:]*\s*(?P<mat>{_SW_MATERIAL_RE})',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:OUTER\s+RING|CENTER(?:ING)?\s+RING|CENTRE\s+RING|O\.?R\.?\b|OR\b)',
    ], text)

    same_ring = re.search(
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+INNER\s+(?:AND|&)\s+OUTER\s+(?:CENTER(?:ING)?\s+)?RING',
        text,
        re.IGNORECASE,
    )
    if same_ring:
        both = _sw_norm_material(same_ring.group('mat'))
        inner = inner or both
        outer = outer or both

    if not inner and winding and re.search(r'\b(?:WITH\s+)?I\s+RING\b|\bINNER\s+RING\b', text, re.IGNORECASE):
        inner = winding
    if not outer:
        proceed_outer = re.search(rf'OUTER\s+RING\s+AS\s+"?(?P<mat>{_SW_MATERIAL_RE})"?', text, re.IGNORECASE)
        if proceed_outer:
            outer = _sw_norm_material(proceed_outer.group('mat'))

    # Domain correction for compact/garbled rows: alloy winding SPW typically
    # uses matching alloy IR and CS OR. Some exports concatenate text as
    # "CS IRAlloy 825 OR"; use the construction pattern rather than the bad
    # boundary when both materials are present after the filler.
    if winding and winding not in ('CS', 'LTCS', 'SS') and re.search(rf'\bCS\s+(?:IR|I\.R\.)\s+(?:{_SW_MATERIAL_RE})\s+(?:OR|O\.R\.)\b', text, re.IGNORECASE):
        inner = winding
        outer = 'CS'

    if inner:
        result['sw_inner_ring'] = inner
    if outer:
        result['sw_outer_ring'] = outer

    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating

    thk_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:NOM\s+)?(?:GASKET\s+)?THK|(\d+(?:\.\d+)?)\s*MM\s+NOM\s+THK', text, re.IGNORECASE)
    if thk_match:
        result['thickness_mm'] = float(thk_match.group(1) or thk_match.group(2))

    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    return result


def _extract_first_size(text: str) -> str | None:
    s = text.upper()
    match = re.search(r'\b(?:NPS|SIZE\s+IN\s+INCH|SIZE|DN)\s*:?\s*(\d+(?:\.\d+)?|\d+\s+\d+/\d+|\d+/\d+|[¼½¾])\s*(?:["\x94]|\'{1,2}|INCH|IN)?\b', s)
    if match:
        return _size_from_text(match.group(1))
    match = re.search(r'\b(\d+(?:\.\d+)?|\d+\s+\d+/\d+|\d+/\d+|[¼½¾])\s*(?:["\x94]|\'{1,2}|INCH|IN)', s)
    if match:
        return _size_from_text(match.group(1))
    match = re.search(r'^\s*(\d+(?:\.\d+)?|[¼½¾])\s*(?:,|GASKET|INSULATING|ISOLATING|INST\.?|IN\b)', s)
    if match:
        return _size_from_text(match.group(1))
    # Concatenated DN/PN exports: "...EN 1514 25PN16..." means DN25 PN16.
    match = re.search(r'\bEN\s*1514\s*(\d{2,4})\s*PN\s*\d{1,2}(?!\d)', s)
    if match:
        return f'{match.group(1)}MM'
    return None


def _parse_number(value: str) -> float:
    return float(value.replace(',', '.'))


def _extract_od_id_thk(text: str) -> tuple[float | None, float | None, float | None, bool]:
    """Return OD, ID, THK, and whether the source was explicitly ID-first."""
    s = text.upper().replace(',', '.')
    od = id_ = thk = None
    id_first = False

    labeled = [
        (r'\bO\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bI\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)', False),
        (r'\bI\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bO\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)', True),
        (r'\bO/?D\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bI/?D\s*[:=]?\s*(\d+(?:\.\d+)?)', False),
        (r'\bI/?D\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bO/?D\s*[:=]?\s*(\d+(?:\.\d+)?)', True),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O\.?D\.?\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I\.?D\.?\b', False),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I\.?D\.?\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O\.?D\.?\b', True),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D\b', False),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\b', True),
        (r'\bINSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM.*?\bOUTSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM', True),
        (r'\bOUTSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM.*?\bINSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM', False),
    ]
    for pattern, reversed_order in labeled:
        match = re.search(pattern, s, re.IGNORECASE)
        if match:
            first, second = _parse_number(match.group(1)), _parse_number(match.group(2))
            if reversed_order:
                id_, od = first, second
                id_first = True
            else:
                od, id_ = first, second
            break

    if od is None or id_ is None:
        # Common custom gasket shorthand: ID x OD x THK.
        match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?', s)
        if match:
            first, second, third = map(_parse_number, match.groups())
            id_, od, thk = first, second, third
            id_first = True

    if od is None or id_ is None:
        match = re.search(r'\bOD\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*ID\s*(\d+(?:\.\d+)?)', s)
        if match:
            od, thk, id_ = map(_parse_number, match.groups())

    if od is None or id_ is None:
        match = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+OD\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s+ID\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s*THK', s)
        if match:
            od, id_, thk = map(_parse_number, match.groups())

    if od is None or id_ is None:
        match = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D(?:\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s*THK)?', s)
        if match:
            od, id_ = _parse_number(match.group(1)), _parse_number(match.group(2))
            if match.group(3):
                thk = _parse_number(match.group(3))

    thk_match = re.search(r'\b(?:THK|THICKNESS)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b|(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:THK|THICK)\b', s)
    if thk_match:
        thk = _parse_number(thk_match.group(1) or thk_match.group(2))
    else:
        # Pattern like "OD 1430 x3x ID 1404".
        mid = re.search(r'\bOD\s*\d+(?:\.\d+)?\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*ID\b', s)
        if mid:
            thk = _parse_number(mid.group(1))

    return od, id_, thk, id_first


def _extract_rtj_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    size = _extract_first_size(text)
    if size:
        result['size'] = size
        result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    ring = re.search(r'\b(?P<ring>R|RX|BX)\s*[- ]?\s*(?P<num>\d{1,4})\b', text, re.IGNORECASE)
    if ring:
        result['ring_no'] = f'{ring.group("ring").upper()}-{ring.group("num")}'

    if re.search(r'\bOCT(?:AGONAL)?\b|TYPE\s*O\b|8[-\s]*SIDED', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'OCTAGONAL'
    elif re.search(r'\bOVAL\b|ELLIPTICAL|TYPE\s*R\b', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'OVAL'
    elif re.search(r'\bBX\b', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'BX'

    material = None
    for pattern in (
        r'(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)\s+(?:OCTAGONAL|OVAL|RING\s+JOINT|RING\s+TYPE|RTJ|R/?J)',
        r'(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)\s+(?:OCTAGONAL|OVAL)\s+RING\s+GASKETS?',
        r'(?:OCTA\s+)?R/?J\s+\d{2,4}#?\s*,?\s*(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)',
        r'(?:MOC|MATERIAL)\s*:?\s*(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)',
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            material = _sw_norm_material(match.group('mat'))
            break
    if material:
        if material == 'SOFT IRON':
            material = 'SOFTIRON'
        if re.search(r'GALVANI[ZS]ED|ZINC\s+PLATED', text, re.IGNORECASE) and 'GALVANISED' not in material:
            material += ' GALVANISED'
        result['moc'] = material
    else:
        fallback = _first_match_material([rf'(?P<mat>{_SW_MATERIAL_RE})'], text)
        if fallback:
            result['moc'] = 'SOFTIRON' if fallback == 'SOFT IRON' else fallback

    hardness = re.search(r'\b(\d{2,3})\s*(?:BHN|HB|HRB|HRBW)\b', text, re.IGNORECASE)
    if hardness:
        result['rtj_hardness_bhn'] = float(hardness.group(1))
    return result


def _extract_kamm_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    od, id_, thk, _ = _extract_od_id_thk(text)
    if od is not None and id_ is not None:
        result.update({'size_type': 'OD_ID', 'od_mm': od, 'id_mm': id_})
    else:
        size = _extract_first_size(text)
        if size:
            result['size'] = size
            result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating
    if thk is not None:
        result['thickness_mm'] = thk
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    # Compact KAMM notation: PRO+FILL+CR: SS316L+GRAPHITE+SS316L
    compact = re.search(r'PRO\s*\+\s*FILL\s*\+\s*(?:CR|CENTER(?:ING)?\s*RING)\s*:?\s*(?P<core>[^+,\s]+(?:\s*316L?)?)\s*\+\s*(?P<surface>[^+,\s]+)(?:\s*\([^)]*\))?\s*\+\s*(?P<ring>[^,\s]+(?:\s*316L?)?)', text, re.IGNORECASE)
    if compact:
        result['kamm_core_material'] = _sw_norm_material(compact.group('core'))
        result['kamm_surface_material'] = _sw_norm_filler(compact.group('surface')) or _sw_norm_material(compact.group('surface'))
        result['sw_outer_ring'] = _sw_norm_material(compact.group('ring'))

    if not result.get('kamm_core_material'):
        core = _first_match_material([
            rf'(?:PROFILE|CAM\s*PROFILE|KAMM?PROFILE|GROOVED\s+PROFILE|GROOVED\s+METAL|CORE|INSERT)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s*(?:KAMM?PROFILE|CAM\s*PROFILE|GROOVED\s+PROFILE|PROFILE\s+GASKET)',
            rf'(?:KAMM?PROFILE|CAM\s*PROFILE|GROOVED\s+PROFILE|GMGC).*?(?P<mat>{_SW_MATERIAL_RE})',
            rf'WITH\s+(?P<mat>{_SW_MATERIAL_RE})\s+AND\s+(?:{_SW_MATERIAL_RE})\s+CENTER(?:ING)?\s+RING',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:GPH|GRAPHITE|PTFE)',
            rf'\b(?P<mat>{_SW_MATERIAL_RE})\s*/\s*(?:{_SW_FILLER_RE})',
        ], text)
        if core:
            result['kamm_core_material'] = core

    if not result.get('kamm_surface_material'):
        surface = _first_match_filler([
            rf'(?:LAYER|LAYERS|FACING|COVERING|FILL(?:ER)?)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_FILLER_RE})',
            rf'(?P<mat>{_SW_FILLER_RE})\s*(?:LAYER|LAYERS|FACING|COVERING|FILLER)',
            rf'(?:{_SW_MATERIAL_RE})\s+(?P<mat>GPH|GRAPHITE|PTFE)',
            rf'(?:{_SW_MATERIAL_RE})\s*/\s*(?P<mat>{_SW_FILLER_RE})',
        ], text)
        if surface:
            result['kamm_surface_material'] = surface

    if not result.get('sw_outer_ring'):
        outer = _first_match_material([
            rf'(?:CENTER(?:ING)?\s+RING|CENTRING\s+RING|OUTER\s+RING|CR)\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:CENTER(?:ING)?\s+RING|CENTRING\s+RING|OUTER\s+RING)',
            rf'\b(?:INR\s+)?(?:{_SW_MATERIAL_RE})\s+(?P<mat>{_SW_MATERIAL_RE})\s+CENTER(?:ING)?\s+RING',
        ], text)
        if outer:
            result['sw_outer_ring'] = outer

    if not result.get('size') and result.get('size_type') != 'OD_ID':
        trailing_size = re.search(r',\s*(\d+(?:\.\d+)?)\s*$', text)
        if trailing_size:
            result['size'] = _size_from_text(trailing_size.group(1))
            result['size_type'] = 'NPS'

    core_thk = re.search(r'\bCORE\s+THK\s*[:=]?\s*(\d+(?:\.\d+)?)\s*MM\b|\((\d+(?:\.\d+)?)\s*MM\s+CORE\s+THK\)', text, re.IGNORECASE)
    if core_thk:
        result['kamm_core_thk'] = float(core_thk.group(1) or core_thk.group(2))
    elif result.get('size_type') == 'OD_ID' and result.get('kamm_core_material'):
        result['kamm_core_thk'] = 4.0

    if re.search(r'\bINTEGRAL\s+(?:OUTER\s+)?RING\b', text, re.IGNORECASE):
        result['kamm_integral_outer_ring'] = True

    if result.get('kamm_core_material'):
        result['sw_winding_material'] = result['kamm_core_material']
    if result.get('kamm_surface_material'):
        result['sw_filler'] = result['kamm_surface_material']
    return result


def _extract_dji_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    od, id_, thk, id_first = _extract_od_id_thk(text)
    if od is not None and id_ is not None:
        result.update({'size_type': 'OD_ID', 'od_mm': od, 'id_mm': id_, 'dji_id_first': id_first})
    if thk is not None:
        result['thickness_mm'] = thk

    jacket = _first_match_material([
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:DOUBLE\s+)?JACKET(?:ED)?',
        rf'(?:JACKET(?:ED)?|MATL\.?|MATERIAL)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
        rf'MATERIAL\s+(?P<mat>{_SW_MATERIAL_RE})\s+AND',
    ], text)
    if jacket:
        result['moc'] = jacket

    filler = _first_match_filler([
        rf'(?:AND|WITH)\s+(?P<mat>{_SW_FILLER_RE})\b',
        rf'(?P<mat>{_SW_FILLER_RE})\s+FILLER',
    ], text)
    if filler:
        result['dji_filler'] = filler

    if re.search(r'\bDRAWING\b|AS\s+PER\s+DRAWING', text, re.IGNORECASE):
        result['special'] = 'AS PER DRAWING'
    if re.search(r'\bRF\b|RAISED\s+FACE', text, re.IGNORECASE):
        result['dji_face_type'] = 'RF'
    elif re.search(r'\bFF\b|FULL\s+FACE', text, re.IGNORECASE):
        result['dji_face_type'] = 'FF'
    return result


def _extract_isk_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    compact_size_rating = re.search(r'^\s*(?P<size>\d{1,2}(?:\.\d+)?)(?P<rating>150|300|600|900|1500|2500)#', text)
    if compact_size_rating:
        result['size'] = _size_from_text(compact_size_rating.group('size'))
        result['size_type'] = 'NPS'
        result['rating'] = f'{compact_size_rating.group("rating")}#'
    size = _extract_first_size(text)
    if size and not result.get('size'):
        result['size'] = size
        result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating and not result.get('rating'):
        result['rating'] = rating
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    if re.search(r'\bTYPE[-\s]?D\b|\bRTJ\b|\bR/?J\b', text, re.IGNORECASE):
        result['gasket_type'] = 'ISK_RTJ'
        result['isk_style'] = 'STYLE-N'
        result['isk_type'] = 'TYPE-D'
    elif re.search(r'\bSTYLE[-\s]?CS\b|\bVCS\b|COMMANDER\s+EXTREME', text, re.IGNORECASE):
        result['isk_style'] = 'STYLE-CS'
        result['isk_type'] = 'TYPE-F'
    elif re.search(r'\bTYPE[-\s]?E\b', text, re.IGNORECASE):
        result['isk_style'] = 'TYPE-E'
        result['isk_type'] = 'TYPE-E'
    elif re.search(r'\bTYPE[-\s]?F\b', text, re.IGNORECASE):
        result['isk_style'] = 'TYPE-F'
        result['isk_type'] = 'TYPE-F'
    elif re.search(r'\bSTYLE[-\s]?N\b', text, re.IGNORECASE):
        result['isk_style'] = 'STYLE-N'

    if re.search(r'\bFF\b|FULL\s+FACE|TYPE[-\s]?E\b', text, re.IGNORECASE):
        result['face_type'] = 'FF'
    elif re.search(r'\bRF\b|RAISED\s+FACE|TYPE[-\s]?F\b|RTJ\b', text, re.IGNORECASE):
        result['face_type'] = 'RF'

    gasket_mat = re.search(r'\b(?:GRE|G[-\s]?10|G[-\s]?11|GLASS\s+REINFORCED\s+EPOXY)[A-Z0-9\s()/-]*', text, re.IGNORECASE)
    if gasket_mat:
        mat = gasket_mat.group(0).strip()
        grade = re.search(r'G[-\s]?(10|11)', mat, re.IGNORECASE)
        result['isk_gasket_material'] = f'GRE G-{grade.group(1)}' if grade else 'GRE'
        result['isk_sleeve_material'] = result['isk_gasket_material']
        result['isk_insulating_washer'] = result['isk_gasket_material']

    core = _first_match_material([
        rf'(?:W/|WITH)?\s*(?P<mat>{_SW_MATERIAL_RE})\s+(?:STEEL\s+)?CORE',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+STEEL\s+CORE',
    ], text)
    if core:
        result['isk_core_material'] = core

    washer = _first_match_material([
        rf'(?P<mat>ZINC\s+PLATED\s+CS|{_SW_MATERIAL_RE})\s+WASHER',
        rf'METALLIC\s+WASHER\s+(?P<mat>ZINC\s+PLATED\s+CS|{_SW_MATERIAL_RE})',
    ], text)
    if washer:
        result['isk_washer_material'] = washer

    if re.search(r'\bPTFE\b|TEFLON', text, re.IGNORECASE):
        result['isk_primary_seal'] = 'PTFE SPRING ENERGISED SEAL' if re.search(r'SPRING|ENERGI[ZS]ED|COMMANDER', text, re.IGNORECASE) else 'PTFE'
    if re.search(r'NON[-\s]?FIRE\s+SAFE|SPRING|ENERGI[ZS]ED', text, re.IGNORECASE):
        result['isk_fire_safety'] = 'NON FIRE SAFE'
    elif re.search(r'\bFIRE\s+SAFE\b', text, re.IGNORECASE):
        result['isk_fire_safety'] = 'FIRE SAFE'

    return result


def _material_from_text(description: str) -> str | None:
    patterns = [
        r'MATERIAL\s+STANDARD\s*:?\s*([^,;]+)',
        r'\bMOC\s*:?\s*([^,;]+)',
        r'\bMATERIAL\s*:?\s*([^,;]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r'\s+LOCATION\s*:.*$', '', value, flags=re.IGNORECASE).strip()
            return value or None
    return None


def _enrich_from_description(item: dict) -> dict:
    desc = item.get('raw_description') or item.get('description') or ''
    if not desc:
        return item
    upper = desc.upper()

    item.setdefault('gasket_type', _infer_gasket_type(desc) or 'SOFT_CUT')

    if not item.get('rating'):
        rating = _rating_from_text(desc)
        if rating:
            item['rating'] = rating

    if not item.get('thickness_mm'):
        thk_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:THK|THICK)', upper)
        if thk_match:
            item['thickness_mm'] = float(thk_match.group(1))

    if not item.get('standard'):
        standard = _standard_from_text(desc)
        if standard:
            item['standard'] = standard

    if not item.get('face_type'):
        if re.search(r'\b(?:FULL\s+FACE|FF)\b', upper):
            item['face_type'] = 'FF'
        elif re.search(r'\bRF\b|\bRAISED\s+FACE\b', upper):
            item['face_type'] = 'RF'

    gasket_type = item.get('gasket_type')

    if gasket_type == 'RTJ':
        extracted = _extract_rtj_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value

    if gasket_type == 'KAMM':
        extracted = _extract_kamm_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value
        if any(item.get(k) for k in ('kamm_core_material', 'kamm_surface_material', 'sw_winding_material')):
            item.pop('moc', None)

    if gasket_type == 'DJI':
        extracted = _extract_dji_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value

    if gasket_type in ('ISK', 'ISK_RTJ'):
        extracted = _extract_isk_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value
        if extracted.get('gasket_type'):
            item['gasket_type'] = extracted['gasket_type']
            gasket_type = item['gasket_type']

    if item.get('gasket_type') == 'SPIRAL_WOUND':
        extracted = _extract_spw_components(desc)
        for key, value in extracted.items():
            if value and (not item.get(key) or (key == 'rating' and '/' in str(value))):
                item[key] = value

        # A generic pipe/spec material column sometimes arrives as `moc`
        # beside a rich SPW description. Once component fields are present,
        # the SPW MOC must be rebuilt from those fields, not from that context.
        if item.get('moc') and any(item.get(k) for k in (
            'sw_winding_material', 'sw_filler', 'sw_inner_ring', 'sw_outer_ring'
        )):
            if not re.search(r'\b(?:WIND|FILL|INNER|OUTER|CENTER|CENTRE|IR|OR|SPW|SPIRAL)\b', str(item['moc']), re.IGNORECASE):
                item.pop('moc', None)

        if not item.get('sw_winding_material'):
            match = re.search(
                r'\b(?:WINDING|MOC\s*:?)\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+WINDING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_winding_material'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_filler'):
            if re.search(r'\b(?:GRAPH(?:ITE|OIL)|GRAFOIL)\b', upper):
                item['sw_filler'] = 'GRAPHITE'
            elif re.search(r'\bPTFE\b', upper):
                item['sw_filler'] = 'PTFE'
        if not item.get('sw_inner_ring'):
            match = re.search(
                r'\bINNER\s+RING\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+INNER\s+RING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_inner_ring'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_outer_ring'):
            match = re.search(
                r'\b(?:OUTER|CENTERING|CENTRE|CENTER)\s+RING\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+(?:OUTER|CENTERING|CENTRE|CENTER)\s+RING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_outer_ring'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_inner_ring') and re.search(r'\bSS\s+INNER\b', upper):
            item['sw_inner_ring'] = 'SS'
        if not item.get('sw_outer_ring') and re.search(r'\bSS\s+INNER\s+AND\s+OUTER\s+CENTERING\s+RING\b', upper):
            item['sw_outer_ring'] = 'SS'

    if item.get('gasket_type') == 'SOFT_CUT' and not item.get('moc'):
        material = _material_from_text(desc)
        if material:
            item['moc'] = material

    if not item.get('size') and item.get('size_type') != 'OD_ID':
        size_match = re.search(r'\bSIZE\s+IN\s+INCH\s*:?\s*(\d+(?:\.\d+)?)\s*"?', upper)
        if size_match:
            item['size'] = _size_from_text(size_match.group(1))
            item['size_type'] = 'NPS'
        else:
            nb_match = re.match(r'\s*(\d+(?:\.\d+)?)\s*MM\b', upper)
            if nb_match:
                item['size'] = f'{nb_match.group(1)}MM'
                item['size_type'] = 'NB'
            else:
                trailing_nb_match = re.search(r'\((\d+(?:\.\d+)?)\s*MM\)\s*$', upper)
                if trailing_nb_match:
                    item['size'] = f'{trailing_nb_match.group(1)}MM'
                    item['size_type'] = 'NB'

    return item


def _description_from_section_row(row: tuple, col_map: dict) -> str | None:
    desc = _cell_str(row, col_map.get('description'))
    if not desc or not _looks_like_gasket(desc):
        return None

    parts = [desc]
    size = _cell_str(row, col_map.get('size_inch'))
    rating = _cell_str(row, col_map.get('rating'))
    thk = _cell_str(row, col_map.get('thickness'))
    od = _cell_str(row, col_map.get('od_mm'))
    id_ = _cell_str(row, col_map.get('id_mm'))

    if size:
        size_rating = size
        if rating:
            size_rating += f' X {rating}'
        if thk:
            size_rating += f' X {thk}MM THK'
        parts.append(size_rating)
    elif od and id_:
        dim = f'OD {od}MM X ID {id_}MM'
        if thk:
            dim += f' X {thk}MM THK'
        parts.append(dim)
    elif thk:
        _append_field(parts, 'THK', thk, 'MM')

    moc_val = _cell_str(row, col_map.get('moc'))
    desc_type = _infer_gasket_type(desc)
    moc_is_technical = bool(moc_val and _looks_like_gasket(moc_val))
    # For rich SPW description rows, generic MATERIAL columns such as
    # LTCS/INCO are pipe/spec context, not gasket MOC. Keep technical
    # MOC cells, because some sheets store the entire SPW construction
    # in MOC and only a short label in DESCRIPTION.
    if moc_val and moc_val.upper() != desc.upper() and (desc_type != 'SPIRAL_WOUND' or moc_is_technical):
        parts.append(f'MOC: {moc_val}')

    _append_field(parts, 'INNER RING WIDTH', _cell_str(row, col_map.get('inner_ring_width')), 'MM')
    _append_field(parts, 'OUTER RING WIDTH', _cell_str(row, col_map.get('outer_ring_width')), 'MM')
    _append_field(parts, 'REMARKS', _cell_str(row, col_map.get('remarks')))

    return ' '.join(parts)


def _parse_description_sections(ws) -> list[dict]:
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_description_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')
        if line_col is None:
            line_col = _infer_line_no_col(data_rows, col_map.get('description'))

        for row in data_rows:
            desc = _description_from_section_row(row, col_map)
            if not desc:
                continue
            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            if qty_col is not None and qty is None and line_no is None:
                continue
            item = {
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'raw_description': desc,
                'quantity': qty,
                'uom': uom,
                'gasket_type': _infer_gasket_type(desc) or 'SOFT_CUT',
            }
            size = _cell_str(row, col_map.get('size_inch'))
            rating = _cell_str(row, col_map.get('rating'))
            thk = _cell_str(row, col_map.get('thickness'))
            moc = _cell_str(row, col_map.get('moc'))
            od = _cell_str(row, col_map.get('od_mm'))
            id_ = _cell_str(row, col_map.get('id_mm'))
            if size:
                item['size'] = _size_from_text(size)
                item['size_type'] = 'NPS'
            if rating:
                item['rating'] = _rating_from_text(rating)
            if thk:
                item['thickness_mm'] = _float_from_text(thk)
            if moc and item['gasket_type'] != 'SPIRAL_WOUND':
                item['moc'] = moc
            if od and id_:
                item['size_type'] = 'OD_ID'
                item['od_mm'] = _float_from_text(od)
                item['id_mm'] = _float_from_text(id_)
            items.append(_enrich_from_description(item))

    return items


def _infer_line_no_col(data_rows: list[tuple], desc_col: int | None) -> int | None:
    """Infer a serial-number column immediately to the left of Description."""
    if desc_col is None or desc_col <= 0:
        return None
    best_col = None
    best_count = 0
    for col_idx in range(desc_col):
        count = 0
        for row in data_rows[:25]:
            if _cell_float(row, col_idx) is not None:
                count += 1
        if count > best_count:
            best_col = col_idx
            best_count = count
    return best_col if best_count >= 1 else None


def _detect_structured_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all structured-format header blocks in a sheet.

    Returns a list of (header_row_idx_0based, col_map, data_rows) tuples where
    col_map maps field names to 0-based column indices and data_rows are the
    value tuples between this header and the next.
    """
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_structured_header(col_map: dict) -> bool:
        has_material = 'material' in col_map
        has_dn_cls = 'dn_size' in col_map and 'class_rating' in col_map
        has_od_id = 'od_mm' in col_map and 'id_mm' in col_map
        has_size_rating = 'size_inch' in col_map and 'rating' in col_map
        return has_material and (has_dn_cls or has_od_id or has_size_rating)

    for row_idx, row in enumerate(all_rows):
        # Try to interpret this row as a header
        col_map = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            if not norm:
                continue
            col_type = _classify_structured_col(norm)
            if col_type and col_type not in col_map:
                col_map[col_type] = col_idx

        if _is_structured_header(col_map):
            # Save previous section if any
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None:
            # Accumulate data rows (skip fully-empty rows)
            if any(c is not None for c in row):
                current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _parse_structured_sheet(ws) -> list[dict]:
    """Parse a sheet that uses one column per field (no combined description column)."""
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_structured_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        mat_col = col_map.get('material')
        dn_col = col_map.get('dn_size')
        cls_col = col_map.get('class_rating')
        thk_col = col_map.get('thickness')
        od_col = col_map.get('od_mm')
        id_col = col_map.get('id_mm')
        size_col = col_map.get('size_inch')
        rating_col = col_map.get('rating')
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')

        last_material = None

        for row in data_rows:
            # Fill-down for MATERIAL column
            mat_val = _cell_str(row, mat_col) if mat_col is not None else None
            if mat_val:
                last_material = mat_val
            else:
                mat_val = last_material

            if not mat_val:
                continue

            # Build synthetic description
            if dn_col is not None and cls_col is not None:
                dn = _cell_str(row, dn_col)
                cls = _cell_str(row, cls_col)
                if not dn or not cls:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{dn} NB {cls}#{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size': f'{dn} NB',
                    'size_type': 'NB',
                    'rating': _rating_from_text(cls),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            elif od_col is not None and id_col is not None:
                od = _cell_str(row, od_col)
                id_ = _cell_str(row, id_col)
                if not od or not id_:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' X {thk}MM THK' if thk else ''
                desc = f'OD {od}MM X ID {id_}MM{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size_type': 'OD_ID',
                    'od_mm': _float_from_text(od),
                    'id_mm': _float_from_text(id_),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            elif size_col is not None and rating_col is not None:
                size_raw = _cell_str(row, size_col)
                rating_raw = _cell_str(row, rating_col)
                if not size_raw or not rating_raw:
                    continue
                # Numeric inch value (e.g. 10 → 10", 1.5 → 1.5")
                try:
                    sv = float(size_raw)
                    size_str = f'{int(sv)}"' if sv == int(sv) else f'{sv}"'
                except ValueError:
                    size_str = size_raw  # already a formatted string
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{size_str} {rating_raw}{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size': size_str,
                    'size_type': 'NPS',
                    'rating': _rating_from_text(rating_raw),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            else:
                continue

            if not _looks_like_gasket(desc):
                continue

            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            items.append(_enrich_from_description({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'raw_description': desc,
                'quantity': qty,
                'uom': uom,
                'gasket_type': _infer_gasket_type(desc) or 'SOFT_CUT',
                **{key: value for key, value in fields.items() if value is not None},
            }))

    return items


# Keywords to identify column types — order matters: more specific first
_HEADER_PATTERNS = {
    'description': ['description', 'dessription', 'desription', 'desc', 'notes'],
    'quantity':    ['qty', 'quantity', 'gross total', 'balance to order', 'balance', 'required qty', 'count'],
    'uom':         ['uom', 'inv uom'],
    'line_no':     ['sl.no', 'sl no', 'sr. no', 'sr no', 'sr no.', 'sno', 'serial', 'sr. no.', 'sr no'],
    'moc':         ['moc', 'base material'],
    'size_inch':   ['nps', 'size'],
    'rating':      ['class /rating', 'class/rating', 'rating', 'class'],
    'thickness':   ['thickness', 'thk', 'thick'],
    'od_mm':       ['gasket od', '(od)', 'od (', 'od mm'],
    'id_mm':       ['gasket id', '(id)', 'id (', 'id mm'],
    'remarks':     ['remarks', 'remark'],
    'inner_ring_width': ['inner ring width'],
    'outer_ring_width': ['outer / center ring  width', 'outer / center ring width', 'outer ring width',
                         'center ring width', 'centering ring width'],
}


def _detect_header(ws) -> tuple[int, dict]:
    """Scan first 15 rows to find the header row and map column indices."""
    best_row, best_col_map, best_score = 0, {}, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        col_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_lower = str(cell).lower().strip()
            for col_type, keywords in _HEADER_PATTERNS.items():
                if col_type not in col_map and any(kw in cell_lower for kw in keywords):
                    col_map[col_type] = col_idx  # 0-based
        score = len(col_map)
        if score >= 2 and 'description' in col_map and score > best_score:
            best_score, best_row, best_col_map = score, row_idx, col_map
    return best_row, best_col_map


def _cell_str(row: tuple, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    # Normalize Shift+Enter newlines within a cell into a single space
    return re.sub(r'[\r\n]+', ' ', str(val)).strip()


def _cell_float(row: tuple, idx: int | None) -> float | None:
    if idx is None or idx >= len(row):
        return None
    try:
        return float(row[idx])
    except (TypeError, ValueError):
        return None
    (r'SS[-\s]*316L|316L\s*SS|316LSS', 'SS316L'),
    (r'SS[-\s]*316|316\s*SS|316SS', 'SS316'),
    (r'SS[-\s]*304L|304L\s*SS|304LSS', 'SS304L'),
    (r'SS[-\s]*304|304\s*SS|304SS', 'SS304'),
